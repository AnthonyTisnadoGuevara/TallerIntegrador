from datetime import datetime, timezone
from io import BytesIO
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field
from typing import Optional
from app.services.supabase_client import supabase
from fastapi import APIRouter, HTTPException, UploadFile, File
import uuid
import os
from app.utils.document_reader import (
    extraer_texto_desde_url,
    extraer_texto_docx,
    extraer_texto_pdf,
    validar_secciones_silabo,
)
from app.utils.excel_reports import crear_reporte_silabos_excel
from app.utils.silabo_extractor import extraer_datos_silabo_desde_texto

router = APIRouter(prefix="/api/silabos", tags=["Sílabos"])


class SilaboCreate(BaseModel):
    semestre_academico: str
    facultad: Optional[str] = None
    programa_estudios: Optional[str] = None
    asignatura: str
    codigo_asignatura: str
    ciclo: int = Field(..., ge=1, le=10)
    modalidad: Optional[str] = None
    creditos: Optional[int] = Field(None, ge=0)
    total_horas_semestrales: Optional[int] = Field(None, ge=0)
    total_horas_semanales: Optional[int] = Field(None, ge=0)
    fecha_inicio: Optional[str] = None
    fecha_culminacion: Optional[str] = None
    duracion_semanas: Optional[int] = Field(None, gt=0)
    docente_responsable: Optional[str] = None
    correo_docente: Optional[str] = None
    archivo_url: Optional[str] = None
    estado: Optional[str] = "pendiente"
    porcentaje_cumplimiento: Optional[float] = 0
    observacion_general: Optional[str] = None


class EstadoUpdate(BaseModel):
    estado: str
    observacion_general: Optional[str] = None

class SilaboUpdate(BaseModel):
    semestre_academico: Optional[str] = None
    facultad: Optional[str] = None
    programa_estudios: Optional[str] = None
    asignatura: Optional[str] = None
    codigo_asignatura: Optional[str] = None
    ciclo: Optional[int] = Field(None, ge=1, le=10)
    modalidad: Optional[str] = None
    creditos: Optional[int] = Field(None, ge=0)
    total_horas_semestrales: Optional[int] = Field(None, ge=0)
    total_horas_semanales: Optional[int] = Field(None, ge=0)
    fecha_inicio: Optional[str] = None
    fecha_culminacion: Optional[str] = None
    duracion_semanas: Optional[int] = Field(None, gt=0)
    docente_responsable: Optional[str] = None
    correo_docente: Optional[str] = None
    archivo_url: Optional[str] = None
    observacion_general: Optional[str] = None

def normalizar_porcentaje_por_estado(silabo: dict) -> dict:
    estado = silabo.get("estado")
    porcentaje = silabo.get("porcentaje_cumplimiento")
    try:
        porcentaje_numero = float(porcentaje or 0)
    except (TypeError, ValueError):
        porcentaje_numero = 0

    if estado == "completo" and (porcentaje is None or porcentaje_numero == 0):
        silabo["porcentaje_cumplimiento"] = 100
    elif estado == "pendiente" and porcentaje is None:
        silabo["porcentaje_cumplimiento"] = 0

    return silabo


def _safe_select_table_reporte_silabos(tabla: str, columnas: str = "*") -> list[dict]:
    try:
        response = supabase.table(tabla).select(columnas).execute()
        return response.data or []
    except Exception as e:
        print(f"[Reporte Sílabos] Advertencia: no se pudo consultar tabla {tabla}.", type(e).__name__)
        return []


def _valor_excel(valor):
    if isinstance(valor, (dict, list)):
        return str(valor)
    if valor is None:
        return ""
    return valor


def _agregar_hoja_excel(workbook, titulo: str, encabezados: list[str], filas: list[list]):
    hoja = workbook.create_sheet(title=titulo[:31])
    hoja.append(encabezados)

    header_fill = PatternFill("solid", fgColor="DCEBFF")
    header_font = Font(bold=True, color="0F172A")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for cell in hoja[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if filas:
        for fila in filas:
            hoja.append([_valor_excel(valor) for valor in fila])
    else:
        hoja.append(["Sin datos disponibles"] + [""] * (len(encabezados) - 1))

    for row in hoja.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in hoja.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        hoja.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 55)

    hoja.freeze_panes = "A2"
    return hoja


def _acciones_silabos(acciones: list[dict]) -> list[dict]:
    return [
        accion
        for accion in acciones
        if accion.get("macroproceso") == "gestion_silabos"
        or accion.get("origen_tipo") == "brecha_curricular"
        or accion.get("silabo_id")
        or accion.get("ciclo")
    ]


def _resumen_reporte_silabos(
    silabos: list[dict],
    analisis: list[dict],
    brechas: list[dict],
    acciones: list[dict],
) -> dict:
    total = len(silabos)
    completos = sum(1 for item in silabos if item.get("estado") == "completo")
    observados = sum(1 for item in silabos if item.get("estado") == "observado")
    pendientes = sum(1 for item in silabos if item.get("estado") == "pendiente")
    incompletos = sum(1 for item in silabos if item.get("estado") == "incompleto")
    cumplimiento = round((completos / total) * 100) if total else 0

    return {
        "fecha_generacion": datetime.now(timezone.utc).isoformat(),
        "total_silabos": total,
        "silabos_completos": completos,
        "silabos_incompletos": incompletos,
        "silabos_observados": observados,
        "silabos_pendientes": pendientes,
        "cumplimiento_promedio": cumplimiento,
        "total_analisis_ia": len(analisis),
        "total_brechas_curriculares": len(brechas),
        "total_acciones_mejora": len(acciones),
    }


def _crear_excel_reporte_silabos(
    resumen: dict,
    silabos: list[dict],
    analisis: list[dict],
    trazabilidad: list[dict],
    brechas: list[dict],
    acciones: list[dict],
) -> BytesIO:
    return crear_reporte_silabos_excel(
        resumen,
        silabos,
        analisis,
        trazabilidad,
        brechas,
        acciones,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    _agregar_hoja_excel(
        workbook,
        "Resumen",
        ["Indicador", "Valor"],
        [
            ["Fecha de generación", resumen.get("fecha_generacion")],
            ["Total de sílabos", resumen.get("total_silabos")],
            ["Sílabos completos", resumen.get("silabos_completos")],
            ["Sílabos incompletos", resumen.get("silabos_incompletos")],
            ["Sílabos observados", resumen.get("silabos_observados")],
            ["Sílabos pendientes", resumen.get("silabos_pendientes")],
            ["Cumplimiento promedio", f"{resumen.get('cumplimiento_promedio', 0)}%"],
            ["Total de análisis IA", resumen.get("total_analisis_ia")],
            ["Total de brechas curriculares", resumen.get("total_brechas_curriculares")],
            ["Total de acciones de mejora", resumen.get("total_acciones_mejora")],
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Sílabos",
        [
            "ID",
            "Semestre académico",
            "Facultad",
            "Programa de estudios",
            "Asignatura",
            "Código",
            "Ciclo",
            "Modalidad",
            "Créditos",
            "Horas semestrales",
            "Horas semanales",
            "Docente",
            "Correo",
            "Estado",
            "Porcentaje de cumplimiento",
            "Observación",
            "Archivo URL",
            "Fecha de registro",
            "Fecha de actualización",
        ],
        [
            [
                item.get("id"),
                item.get("semestre_academico"),
                item.get("facultad"),
                item.get("programa_estudios"),
                item.get("asignatura"),
                item.get("codigo_asignatura"),
                item.get("ciclo"),
                item.get("modalidad"),
                item.get("creditos"),
                item.get("total_horas_semestrales"),
                item.get("total_horas_semanales"),
                item.get("docente_responsable"),
                item.get("correo_docente"),
                item.get("estado"),
                item.get("porcentaje_cumplimiento"),
                item.get("observacion_general"),
                item.get("archivo_url"),
                item.get("created_at"),
                item.get("updated_at"),
            ]
            for item in silabos
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Análisis IA",
        ["ID", "Sílabo ID", "Asignatura", "Modelo usado", "Nivel de cumplimiento", "Resumen", "Recomendaciones", "Fecha de análisis"],
        [
            [
                item.get("id"),
                item.get("silabo_id"),
                item.get("asignatura"),
                item.get("modelo_usado"),
                item.get("nivel_cumplimiento") or item.get("nivel_riesgo"),
                item.get("resumen"),
                item.get("recomendaciones") or item.get("sugerencias"),
                item.get("created_at"),
            ]
            for item in analisis
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Trazabilidad Curricular",
        ["ID", "Curso origen", "Curso destino", "Ciclo origen", "Ciclo destino", "Tipo de relación", "Nivel de coherencia", "Observación", "Recomendación", "Fecha"],
        [
            [
                item.get("id"),
                item.get("asignatura_origen") or item.get("curso_origen"),
                item.get("asignatura_destino") or item.get("curso_destino"),
                item.get("ciclo_origen"),
                item.get("ciclo_destino"),
                item.get("tipo_relacion"),
                item.get("nivel_coherencia"),
                item.get("observacion"),
                item.get("sugerencia") or item.get("recomendacion"),
                item.get("created_at"),
            ]
            for item in trazabilidad
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Brechas Curriculares",
        ["ID", "Tipo de brecha", "Descripción", "Prioridad", "Curso relacionado", "Ciclo", "Recomendación", "Estado", "Fecha"],
        [
            [
                item.get("id"),
                item.get("tipo_brecha"),
                item.get("descripcion"),
                item.get("prioridad"),
                item.get("asignatura") or item.get("curso_relacionado"),
                item.get("ciclo"),
                item.get("recomendacion"),
                item.get("estado"),
                item.get("created_at"),
            ]
            for item in brechas
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Acciones de Mejora",
        ["ID", "Título", "Descripción", "Prioridad", "Estado", "Responsable", "Origen", "Fecha programada", "Fecha de creación"],
        [
            [
                item.get("id"),
                item.get("titulo"),
                item.get("descripcion"),
                item.get("prioridad"),
                item.get("estado"),
                item.get("responsable"),
                item.get("origen_tipo"),
                item.get("fecha_limite") or item.get("fecha_programada"),
                item.get("created_at"),
            ]
            for item in acciones
        ],
    )

    archivo = BytesIO()
    workbook.save(archivo)
    archivo.seek(0)
    return archivo


@router.get("/")
def listar_silabos():
    try:
        response = supabase.table("silabos").select("*").order("ciclo").execute()
        silabos = [normalizar_porcentaje_por_estado(item) for item in (response.data or [])]
        return {
            "message": "Listado de sílabos obtenido correctamente",
            "data": silabos
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/extraer-datos-archivo")
async def extraer_datos_archivo_silabo(archivo: UploadFile = File(...)):
    try:
        nombre_original = archivo.filename or ""
        extension = os.path.splitext(nombre_original)[1].lower()

        if extension not in [".docx", ".pdf"]:
            raise HTTPException(
                status_code=400,
                detail="Formato no permitido. Solo se aceptan archivos DOCX o PDF."
            )

        contenido = await archivo.read()
        if not contenido:
            raise HTTPException(status_code=400, detail="El archivo esta vacio.")

        try:
            if extension == ".docx":
                texto = extraer_texto_docx(contenido)
            else:
                texto = extraer_texto_pdf(contenido)
        except ValueError as error:
            mensaje_error = str(error)
            if "no contiene texto" in mensaje_error.lower() or "texto extra" in mensaje_error.lower():
                raise HTTPException(
                    status_code=400,
                    detail="No se pudo extraer texto del documento. Si es PDF escaneado, conviertalo a PDF con texto seleccionable o a DOCX."
                ) from error
            raise HTTPException(status_code=400, detail=mensaje_error) from error

        if not texto.strip():
            raise HTTPException(
                status_code=400,
                detail="No se pudo extraer texto del documento. Si es PDF escaneado, conviertalo a PDF con texto seleccionable o a DOCX."
            )

        datos = extraer_datos_silabo_desde_texto(texto)

        return {
            "message": "Datos extraidos correctamente",
            "data": datos,
            "preview_texto": texto[:500]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reporte/excel")
def exportar_reporte_silabos_excel():
    try:
        print("[Reporte Sílabos] Generando reporte Excel...")

        silabos = [
            normalizar_porcentaje_por_estado(item)
            for item in _safe_select_table_reporte_silabos("silabos")
        ]
        analisis = _safe_select_table_reporte_silabos("analisis_silabo")
        trazabilidad = _safe_select_table_reporte_silabos("trazabilidad_curricular")
        brechas = _safe_select_table_reporte_silabos("brechas_curriculares")
        acciones = _acciones_silabos(_safe_select_table_reporte_silabos("acciones_mejora"))
        resumen = _resumen_reporte_silabos(silabos, analisis, brechas, acciones)
        archivo = _crear_excel_reporte_silabos(
            resumen,
            silabos,
            analisis,
            trazabilidad,
            brechas,
            acciones,
        )

        print("[Reporte Sílabos] Reporte generado correctamente.")
        return StreamingResponse(
            archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="reporte_gestion_silabos.xlsx"'
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/{silabo_id}")
def obtener_silabo(silabo_id: str):
    try:
        response = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not response.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        return {
            "message": "Sílabo obtenido correctamente",
            "data": response.data[0]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/")
def crear_silabo(silabo: SilaboCreate):
    try:
        estados_validos = ["pendiente", "completo", "observado", "incompleto"]

        if silabo.estado not in estados_validos:
            raise HTTPException(
                status_code=400,
                detail="Estado inválido. Use: pendiente, completo, observado o incompleto"
            )

        if silabo.ciclo < 1 or silabo.ciclo > 10:
            raise HTTPException(
                status_code=400,
                detail="El ciclo debe estar entre 1 y 10"
            )

        datos_silabo = silabo.model_dump()
        if datos_silabo["estado"] == "completo":
            datos_silabo["porcentaje_cumplimiento"] = 100
        elif datos_silabo["estado"] == "pendiente":
            datos_silabo["porcentaje_cumplimiento"] = 0

        response = (
            supabase.table("silabos")
            .insert(datos_silabo)
            .execute()
        )

        return {
            "message": "Sílabo registrado correctamente",
            "data": response.data[0] if response.data else None
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{silabo_id}/estado")
def actualizar_estado_silabo(silabo_id: str, datos: EstadoUpdate):
    try:
        estados_validos = ["pendiente", "completo", "observado", "incompleto"]

        if datos.estado not in estados_validos:
            raise HTTPException(
                status_code=400,
                detail="Estado inválido. Use: pendiente, completo, observado o incompleto"
            )

        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        estado_anterior = silabo_actual.data[0]["estado"]
        datos_actualizar = {
            "estado": datos.estado,
            "observacion_general": datos.observacion_general
        }

        if datos.estado == "completo":
            datos_actualizar["porcentaje_cumplimiento"] = 100
        elif datos.estado == "pendiente":
            datos_actualizar["porcentaje_cumplimiento"] = 0

        response = (
            supabase.table("silabos")
            .update(datos_actualizar)
            .eq("id", silabo_id)
            .execute()
        )

        supabase.table("historial_silabo").insert({
            "silabo_id": silabo_id,
            "estado_anterior": estado_anterior,
            "estado_nuevo": datos.estado,
            "observacion": datos.observacion_general,
            "usuario": "backend"
        }).execute()

        return {
            "message": "Estado del sílabo actualizado correctamente",
            "data": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{silabo_id}")
def eliminar_silabo(silabo_id: str):
    try:
        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        response = (
            supabase.table("silabos")
            .delete()
            .eq("id", silabo_id)
            .execute()
        )

        return {
            "message": "Sílabo eliminado correctamente",
            "data": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/estado/alertas")
def listar_silabos_con_alertas():
    try:
        response = (
            supabase.table("silabos")
            .select("*")
            .in_("estado", ["observado", "incompleto", "pendiente"])
            .order("ciclo")
            .execute()
        )

        return {
            "message": "Sílabos con alertas obtenidos correctamente",
            "data": response.data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@router.get("/{silabo_id}/historial")
def obtener_historial_silabo(silabo_id: str):
    try:
        # Verificar si el sílabo existe
        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        # Consultar historial del sílabo
        response = (
            supabase.table("historial_silabo")
            .select("*")
            .eq("silabo_id", silabo_id)
            .order("created_at", desc=True)
            .execute()
        )

        return {
            "message": "Historial del sílabo obtenido correctamente",
            "silabo": {
                "id": silabo_actual.data[0]["id"],
                "asignatura": silabo_actual.data[0]["asignatura"],
                "codigo_asignatura": silabo_actual.data[0]["codigo_asignatura"],
                "estado_actual": silabo_actual.data[0]["estado"]
            },
            "historial": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/{silabo_id}/archivo")
async def subir_archivo_silabo(silabo_id: str, archivo: UploadFile = File(...)):
    try:
        # Verificar si el sílabo existe
        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        # Validar extensión del archivo
        nombre_original = archivo.filename
        extension = os.path.splitext(nombre_original)[1].lower()

        extensiones_permitidas = [".pdf", ".docx"]

        if extension not in extensiones_permitidas:
            raise HTTPException(
                status_code=400,
                detail="Formato no permitido. Solo se aceptan archivos PDF o DOCX."
            )

        # Leer contenido del archivo
        contenido = await archivo.read()

        # Crear nombre único para evitar duplicados
        nombre_archivo = f"{silabo_id}_{uuid.uuid4()}{extension}"

        # Ruta dentro del bucket
        ruta_storage = f"silabos/{nombre_archivo}"

        # Subir archivo a Supabase Storage
        supabase.storage.from_("silabos").upload(
            path=ruta_storage,
            file=contenido,
            file_options={
                "content-type": archivo.content_type or "application/octet-stream",
                "upsert": "true"
            }
        )

        # Obtener URL pública del archivo
        archivo_url = supabase.storage.from_("silabos").get_public_url(ruta_storage)

        # Actualizar campo archivo_url en tabla silabos
        response = (
            supabase.table("silabos")
            .update({
                "archivo_url": archivo_url,
                "updated_at": datetime.now(timezone.utc).isoformat()
            })
            .eq("id", silabo_id)
            .execute()
        )

        return {
            "message": "Archivo del silabo actualizado correctamente.",
            "archivo_url": archivo_url,
            "data": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/{silabo_id}/validar-documento")
def validar_documento_silabo(silabo_id: str):
    try:
        # Verificar si el sílabo existe
        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        silabo = silabo_actual.data[0]
        archivo_url = silabo.get("archivo_url")

        if not archivo_url:
            raise HTTPException(
                status_code=400,
                detail="El sílabo no tiene archivo cargado para validar."
            )

        archivo_url_limpia = archivo_url.lower().split("?")[0]
        if archivo_url_limpia.endswith(".doc"):
            raise HTTPException(
                status_code=400,
                detail="El formato .doc antiguo no está soportado. Convierta el archivo a .docx o PDF con texto seleccionable."
            )

        # Extraer texto del documento DOCX o PDF mediante el lector central
        try:
            texto = extraer_texto_desde_url(archivo_url)
        except ValueError as error:
            mensaje_error = str(error)
            if "PDF no contiene texto extra" in mensaje_error:
                raise HTTPException(
                    status_code=400,
                    detail="No se pudo extraer texto del documento. Si es PDF escaneado, conviértalo a PDF con texto seleccionable o a DOCX."
                ) from error
            if "formato .doc antiguo" in mensaje_error.lower():
                raise HTTPException(
                    status_code=400,
                    detail="El formato .doc antiguo no está soportado. Convierta el archivo a .docx o PDF con texto seleccionable."
                ) from error
            raise HTTPException(
                status_code=400,
                detail="Solo se permite validar documentos .docx o PDF con texto extraíble."
            ) from error
        except RuntimeError as error:
            raise HTTPException(status_code=400, detail=str(error)) from error

        if not texto.strip():
            raise HTTPException(
                status_code=400,
                detail="No se pudo extraer texto del documento. Si es PDF escaneado, conviértalo a PDF con texto seleccionable o a DOCX."
            )

        # Validar secciones obligatorias
        resultados = validar_secciones_silabo(texto)

        # Eliminar validaciones anteriores del mismo sílabo
        supabase.table("validacion_silabo").delete().eq("silabo_id", silabo_id).execute()

        # Insertar nuevas validaciones
        registros_validacion = [
            {
                "silabo_id": silabo_id,
                "seccion": item["seccion"],
                "cumple": item["cumple"],
                "observacion": item["observacion"]
            }
            for item in resultados
        ]

        supabase.table("validacion_silabo").insert(registros_validacion).execute()

        total = len(resultados)
        cumplidas = sum(1 for item in resultados if item["cumple"])
        porcentaje = round((cumplidas / total) * 100, 2)

        if porcentaje >= 90:
            nuevo_estado = "completo"
        elif porcentaje >= 70:
            nuevo_estado = "observado"
        else:
            nuevo_estado = "incompleto"

        estado_anterior = silabo.get("estado")

        # Actualizar sílabo
        response = (
            supabase.table("silabos")
            .update({
                "porcentaje_cumplimiento": porcentaje,
                "estado": nuevo_estado,
                "observacion_general": f"Validación automática: {cumplidas} de {total} secciones identificadas."
            })
            .eq("id", silabo_id)
            .execute()
        )

        # Registrar historial
        supabase.table("historial_silabo").insert({
            "silabo_id": silabo_id,
            "estado_anterior": estado_anterior,
            "estado_nuevo": nuevo_estado,
            "observacion": f"Validación automática del documento: {cumplidas} de {total} secciones identificadas.",
            "usuario": "backend"
        }).execute()

        return {
            "message": "Documento del sílabo validado correctamente",
            "silabo_id": silabo_id,
            "secciones_cumplidas": cumplidas,
            "total_secciones": total,
            "porcentaje_cumplimiento": porcentaje,
            "estado": nuevo_estado,
            "validacion": resultados,
            "data": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@router.get("/{silabo_id}/validacion")
def obtener_validacion_silabo(silabo_id: str):
    try:
        # Verificar si el sílabo existe
        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        silabo = silabo_actual.data[0]

        # Obtener validaciones del sílabo
        response = (
            supabase.table("validacion_silabo")
            .select("*")
            .eq("silabo_id", silabo_id)
            .order("seccion")
            .execute()
        )

        return {
            "message": "Validación del sílabo obtenida correctamente",
            "silabo": {
                "id": silabo["id"],
                "asignatura": silabo["asignatura"],
                "codigo_asignatura": silabo["codigo_asignatura"],
                "estado": silabo["estado"],
                "porcentaje_cumplimiento": silabo["porcentaje_cumplimiento"],
                "archivo_url": silabo["archivo_url"]
            },
            "validacion": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@router.put("/{silabo_id}")
def actualizar_silabo(silabo_id: str, datos: SilaboUpdate):
    try:
        # Verificar si el sílabo existe
        silabo_actual = (
            supabase.table("silabos")
            .select("*")
            .eq("id", silabo_id)
            .execute()
        )

        if not silabo_actual.data:
            raise HTTPException(status_code=404, detail="Sílabo no encontrado")

        datos_actualizar = datos.model_dump(exclude_unset=True)

        if not datos_actualizar:
            raise HTTPException(
                status_code=400,
                detail="No se enviaron datos para actualizar"
            )

        if "ciclo" in datos_actualizar:
            if datos_actualizar["ciclo"] < 1 or datos_actualizar["ciclo"] > 10:
                raise HTTPException(
                    status_code=400,
                    detail="El ciclo debe estar entre 1 y 10"
                )

        response = (
            supabase.table("silabos")
            .update(datos_actualizar)
            .eq("id", silabo_id)
            .execute()
        )

        # Registrar trazabilidad de edición
        supabase.table("historial_silabo").insert({
            "silabo_id": silabo_id,
            "estado_anterior": silabo_actual.data[0]["estado"],
            "estado_nuevo": silabo_actual.data[0]["estado"],
            "observacion": "Se actualizó información general del sílabo.",
            "usuario": "frontend"
        }).execute()

        return {
            "message": "Información del sílabo actualizada correctamente",
            "data": response.data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
