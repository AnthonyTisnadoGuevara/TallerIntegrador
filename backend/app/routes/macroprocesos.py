from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import os
import uuid
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field

from app.agents.gestion_academica_graph import ejecutar_grafo_gestion_academica
from app.agents.mejora_continua_coordinator import ejecutar_grafo_coordinador_mejora_continua
from app.agents.planificacion_graph import ejecutar_grafo_planificacion
from app.agents.validacion_evidencia_graph import ejecutar_grafo_validacion_evidencia
from app.services.supabase_client import supabase
from app.utils.excel_reports import crear_reporte_integral_excel


router = APIRouter(
    prefix="/api/macroprocesos",
    tags=["Macroprocesos"],
)

ESTADOS_VALIDOS = {"pendiente", "en_proceso", "completado", "observado"}
PRIORIDADES_VALIDAS = {"alta", "media", "baja"}
NIVELES_ALERTA_VALIDOS = {"baja", "media", "alta", "critica"}
ESTADOS_ALERTA_VALIDOS = {"activa", "atendida", "descartada"}
ORDEN_PRIORIDAD = {"alta": 0, "media": 1, "baja": 2}
ORDEN_ESTADO = {"pendiente": 0, "en_proceso": 1, "observado": 2, "completado": 3}
ORDEN_ALERTA = {"critica": 0, "alta": 1, "media": 2, "baja": 3}
MACROPROCESOS_ACCIONES_VALIDOS = {"planificacion_estrategica", "gestion_academica"}
MACROPROCESOS_EVIDENCIAS_VALIDOS = {"planificacion_estrategica", "gestion_academica"}
MACROPROCESOS_ANALISIS_VALIDOS = {"planificacion_estrategica", "gestion_academica", "mejora_continua"}
MACROPROCESOS_SEMAFORO = {
    "planificacion_estrategica": "Planificación Estratégica",
    "gestion_academica": "Gestión Académica",
    "gestion_silabos": "Gestión de Sílabos",
}
TIPOS_ANALISIS_VALIDOS = {
    "analisis_planificacion",
    "analisis_gestion_academica",
    "analisis_integral_mejora_continua",
}
ORDEN_RIESGO = {"bajo": 1, "medio": 2, "alto": 3}
CAMPOS_HISTORIAL = {
    "estado",
    "prioridad",
    "avance",
    "observacion",
    "archivo_url",
    "fecha_cumplimiento",
    "responsable",
    "descripcion",
    "tipo_evidencia",
}
EXTENSIONES_EVIDENCIA = {".pdf", ".docx", ".doc", ".xlsx", ".png", ".jpg", ".jpeg"}
TIPOS_CONTENIDO_EVIDENCIA = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "image/png",
    "image/jpeg",
}
BUCKET_EVIDENCIAS = os.getenv("SUPABASE_STORAGE_BUCKET", "evidencias").strip() or "evidencias"
NIVELES_AVANCE_SEMANAL = {"Sin avance", "Bajo", "Medio", "Alto", "Completado"}
TIPOS_APOYO_SEMANAL = {"Directivo", "Académico", "Documental", "Tecnológico", "Otro"}


class EvidenciaCreate(BaseModel):
    macroproceso: str
    codigo: Optional[str] = None
    titulo: str
    descripcion: Optional[str] = None
    tipo_evidencia: Optional[str] = None
    responsable: Optional[str] = None
    mes_programado: Optional[str] = None
    fecha_programada: Optional[str] = None
    fecha_cumplimiento: Optional[str] = None
    estado: Optional[str] = "pendiente"
    prioridad: Optional[str] = "media"
    avance: Optional[int] = Field(0, ge=0, le=100)
    observacion: Optional[str] = None
    archivo_url: Optional[str] = None
    origen_documento: Optional[str] = None


class EvidenciaUpdate(BaseModel):
    macroproceso: Optional[str] = None
    codigo: Optional[str] = None
    titulo: Optional[str] = None
    descripcion: Optional[str] = None
    tipo_evidencia: Optional[str] = None
    responsable: Optional[str] = None
    mes_programado: Optional[str] = None
    fecha_programada: Optional[str] = None
    fecha_cumplimiento: Optional[str] = None
    estado: Optional[str] = None
    prioridad: Optional[str] = None
    avance: Optional[int] = Field(None, ge=0, le=100)
    observacion: Optional[str] = None
    archivo_url: Optional[str] = None
    origen_documento: Optional[str] = None


class SeguimientoSemanalCreate(BaseModel):
    macroproceso: Optional[str] = None
    codigo_evidencia: Optional[str] = None
    semana_inicio: str
    semana_fin: str
    responsable: Optional[str] = None
    accion_realizada: bool = False
    nivel_avance: str = "Sin avance"
    porcentaje_avance: int = Field(0, ge=0, le=100)
    descripcion_accion: Optional[str] = None
    resultado_observado: Optional[str] = None
    dificultad_encontrada: Optional[str] = None
    compromiso_siguiente_semana: Optional[str] = None
    requiere_apoyo: bool = False
    tipo_apoyo_requerido: Optional[str] = None
    observacion: Optional[str] = None


class SeguimientoSemanalUpdate(BaseModel):
    semana_inicio: Optional[str] = None
    semana_fin: Optional[str] = None
    responsable: Optional[str] = None
    accion_realizada: Optional[bool] = None
    nivel_avance: Optional[str] = None
    porcentaje_avance: Optional[int] = Field(None, ge=0, le=100)
    descripcion_accion: Optional[str] = None
    resultado_observado: Optional[str] = None
    dificultad_encontrada: Optional[str] = None
    compromiso_siguiente_semana: Optional[str] = None
    requiere_apoyo: Optional[bool] = None
    tipo_apoyo_requerido: Optional[str] = None
    observacion: Optional[str] = None


class AlertaUpdate(BaseModel):
    estado: Optional[str] = None
    recomendacion: Optional[str] = None


class EncuestaAceptacionCreate(BaseModel):
    nombre_usuario: Optional[str] = None
    rol_usuario: Optional[str] = None
    claridad_interfaz: int = Field(..., ge=1, le=5)
    facilidad_uso: int = Field(..., ge=1, le=5)
    utilidad_evidencias: int = Field(..., ge=1, le=5)
    utilidad_ia: int = Field(..., ge=1, le=5)
    utilidad_alertas: int = Field(..., ge=1, le=5)
    utilidad_semaforo: int = Field(..., ge=1, le=5)
    satisfaccion_general: int = Field(..., ge=1, le=5)
    comentario: Optional[str] = None


def _ahora_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalizar_estado(estado: Optional[str]) -> Optional[str]:
    if estado is None:
        return None

    estado_normalizado = estado.strip().lower()
    if estado_normalizado not in ESTADOS_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="El estado debe ser: pendiente, en_proceso, completado u observado.",
        )
    return estado_normalizado


def _normalizar_prioridad(prioridad: Optional[str]) -> Optional[str]:
    if prioridad is None:
        return None

    prioridad_normalizada = prioridad.strip().lower()
    if prioridad_normalizada not in PRIORIDADES_VALIDAS:
        raise HTTPException(
            status_code=400,
            detail="La prioridad debe ser: alta, media o baja.",
        )
    return prioridad_normalizada


def _normalizar_nivel_alerta(nivel_alerta: Optional[str]) -> Optional[str]:
    if nivel_alerta is None:
        return None

    nivel_normalizado = nivel_alerta.strip().lower()
    if nivel_normalizado not in NIVELES_ALERTA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="El nivel de alerta debe ser: baja, media, alta o critica.",
        )
    return nivel_normalizado


def _normalizar_estado_alerta(estado: Optional[str]) -> Optional[str]:
    if estado is None:
        return None

    estado_normalizado = estado.strip().lower()
    if estado_normalizado not in ESTADOS_ALERTA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="El estado de alerta debe ser: activa, atendida o descartada.",
        )
    return estado_normalizado


def _normalizar_payload(data: dict) -> dict:
    if "estado" in data:
        data["estado"] = _normalizar_estado(data.get("estado"))
    if "prioridad" in data:
        data["prioridad"] = _normalizar_prioridad(data.get("prioridad"))
    if "macroproceso" in data and data.get("macroproceso"):
        data["macroproceso"] = data["macroproceso"].strip().lower()
    return data


def _parsear_fecha_iso(valor: str, campo: str) -> date:
    try:
        return date.fromisoformat(valor)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"{campo} debe tener formato YYYY-MM-DD.") from error


def _validar_payload_seguimiento(data: dict) -> None:
    inicio = _parsear_fecha_iso(data["semana_inicio"], "semana_inicio")
    fin = _parsear_fecha_iso(data["semana_fin"], "semana_fin")
    if inicio > fin:
        raise HTTPException(status_code=400, detail="semana_inicio no debe ser mayor que semana_fin.")

    if data.get("nivel_avance") not in NIVELES_AVANCE_SEMANAL:
        raise HTTPException(
            status_code=400,
            detail="nivel_avance debe ser: Sin avance, Bajo, Medio, Alto o Completado.",
        )

    if data.get("requiere_apoyo") and not data.get("tipo_apoyo_requerido"):
        raise HTTPException(status_code=400, detail="tipo_apoyo_requerido es obligatorio si requiere_apoyo es true.")

    if data.get("tipo_apoyo_requerido") and data.get("tipo_apoyo_requerido") not in TIPOS_APOYO_SEMANAL:
        raise HTTPException(
            status_code=400,
            detail="tipo_apoyo_requerido debe ser: Directivo, Académico, Documental, Tecnológico u Otro.",
        )

    if data.get("accion_realizada") is False and not (
        data.get("compromiso_siguiente_semana") or data.get("dificultad_encontrada")
    ):
        raise HTTPException(
            status_code=400,
            detail="Si no se realizó acción, registre una dificultad o compromiso para la siguiente semana.",
        )


def _obtener_seguimiento_o_404(seguimiento_id: str) -> dict:
    response = (
        supabase.table("seguimiento_semanal_evidencias")
        .select("*")
        .eq("id", seguimiento_id)
        .execute()
    )
    if not response.data:
        raise HTTPException(status_code=404, detail="Seguimiento semanal no encontrado.")
    return response.data[0]


def _ordenar_evidencias(evidencias: list[dict]) -> list[dict]:
    return sorted(
        evidencias,
        key=lambda item: (
            item.get("macroproceso") or "",
            ORDEN_PRIORIDAD.get(item.get("prioridad"), 99),
            ORDEN_ESTADO.get(item.get("estado"), 99),
            item.get("codigo") or "",
        ),
    )


def _resumen_evidencias(evidencias: list[dict]) -> dict:
    total = len(evidencias)
    avance_promedio = 0
    if total:
        avance_promedio = round(
            sum(int(item.get("avance") or 0) for item in evidencias) / total
        )

    return {
        "total": total,
        "pendientes": sum(1 for item in evidencias if item.get("estado") == "pendiente"),
        "en_proceso": sum(1 for item in evidencias if item.get("estado") == "en_proceso"),
        "completadas": sum(1 for item in evidencias if item.get("estado") == "completado"),
        "observadas": sum(1 for item in evidencias if item.get("estado") == "observado"),
        "avance_promedio": avance_promedio,
        "alta": sum(1 for item in evidencias if item.get("prioridad") == "alta"),
        "media": sum(1 for item in evidencias if item.get("prioridad") == "media"),
        "baja": sum(1 for item in evidencias if item.get("prioridad") == "baja"),
    }


def _obtener_evidencia_o_404(evidencia_id: str) -> dict:
    response = (
        supabase.table("macroproceso_evidencias")
        .select("*")
        .eq("id", evidencia_id)
        .execute()
    )
    if not response.data:
        raise HTTPException(status_code=404, detail="Evidencia no encontrada.")
    return response.data[0]


def _safe_select_table(tabla: str, columnas: str = "*") -> list[dict]:
    try:
        response = supabase.table(tabla).select(columnas).execute()
        return response.data or []
    except Exception as e:
        print(f"[Alertas macroprocesos] No se pudo leer {tabla}:", type(e).__name__, str(e))
        return []


def _safe_select_table_reporte(tabla: str, columnas: str = "*") -> list[dict]:
    try:
        response = supabase.table(tabla).select(columnas).execute()
        return response.data or []
    except Exception as e:
        print(f"[Reporte Integral] Advertencia: no se pudo consultar tabla {tabla}.", type(e).__name__)
        return []


def _valor_excel(valor):
    if isinstance(valor, (dict, list)):
        return str(valor)
    if valor is None:
        return ""
    return valor


def _agregar_hoja_excel(workbook, titulo: str, encabezados: list[str], filas: list[list]):
    hoja = workbook.create_sheet(title=titulo[:31])
    hoja.append(encabezados)

    header_fill = PatternFill("solid", fgColor="DCEBFF")
    header_font = Font(bold=True, color="0F172A")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for cell in hoja[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for fila in filas:
        hoja.append([_valor_excel(valor) for valor in fila])

    for row in hoja.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in hoja.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        hoja.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 55)

    hoja.freeze_panes = "A2"
    return hoja


def _crear_excel_reporte_integral(reporte: dict, metricas: dict) -> BytesIO:
    return crear_reporte_integral_excel(reporte, metricas)

    workbook = Workbook()
    workbook.remove(workbook.active)

    resumen = reporte.get("resumen", {})
    _agregar_hoja_excel(
        workbook,
        "Resumen General",
        ["Indicador", "Valor"],
        [
            ["Título", reporte.get("titulo")],
            ["Fecha de generación", reporte.get("fecha_generacion")],
            ["Total macroprocesos", resumen.get("total_macroprocesos")],
            ["Total evidencias", resumen.get("total_evidencias")],
            ["Alertas activas", resumen.get("total_alertas_activas")],
            ["Alertas críticas", resumen.get("total_alertas_criticas")],
            ["Acciones de mejora", resumen.get("total_acciones_mejora")],
            ["Acciones pendientes", resumen.get("acciones_pendientes")],
            ["Acciones en proceso", resumen.get("acciones_en_proceso")],
            ["Acciones completadas", resumen.get("acciones_completadas")],
            ["Análisis IA", resumen.get("total_analisis_ia")],
            ["Validaciones documentales IA", resumen.get("total_validaciones_ia")],
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Semáforo",
        ["Macroproceso", "Color", "Avance", "Alertas activas", "Nivel de riesgo", "Estado"],
        [
            [
                item.get("macroproceso"),
                item.get("color"),
                item.get("avance_promedio"),
                item.get("alertas_activas"),
                item.get("nivel_riesgo"),
                item.get("estado"),
            ]
            for item in reporte.get("semaforo", [])
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Evidencias",
        ["Macroproceso", "Código", "Título", "Responsable", "Estado", "Prioridad", "Avance", "Archivo", "Observación"],
        [
            [
                item.get("macroproceso"),
                item.get("codigo"),
                item.get("titulo"),
                item.get("responsable"),
                item.get("estado"),
                item.get("prioridad"),
                item.get("avance"),
                item.get("archivo_url"),
                item.get("observacion"),
            ]
            for item in reporte.get("evidencias", reporte.get("evidencias_criticas", []))
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Alertas Activas",
        ["Macroproceso", "Código", "Título", "Nivel", "Estado", "Descripción", "Recomendación"],
        [
            [
                item.get("macroproceso"),
                item.get("codigo"),
                item.get("titulo"),
                item.get("nivel_alerta"),
                item.get("estado"),
                item.get("descripcion"),
                item.get("recomendacion"),
            ]
            for item in reporte.get("alertas_activas", [])
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Acciones de Mejora",
        ["Macroproceso", "Título", "Prioridad", "Estado", "Responsable", "Origen", "Descripción", "Recomendación"],
        [
            [
                item.get("macroproceso"),
                item.get("titulo"),
                item.get("prioridad"),
                item.get("estado"),
                item.get("responsable"),
                item.get("origen_tipo"),
                item.get("descripcion"),
                item.get("recomendacion"),
            ]
            for item in reporte.get("acciones_mejora", [])
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Análisis IA",
        ["Macroproceso", "Tipo", "Nivel de riesgo", "Resumen", "Modelo", "Fecha"],
        [
            [
                item.get("macroproceso"),
                item.get("tipo_analisis"),
                item.get("nivel_riesgo"),
                item.get("resumen"),
                item.get("modelo_usado"),
                item.get("created_at"),
            ]
            for item in reporte.get("ultimos_analisis_ia", [])
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Validaciones Documentales",
        ["Macroproceso", "Evidencia ID", "Nivel", "Pertinencia", "Resumen", "Acción sugerida", "Modelo", "Fecha"],
        [
            [
                item.get("macroproceso"),
                item.get("evidencia_id"),
                item.get("nivel_validez"),
                item.get("pertinencia"),
                item.get("resumen"),
                item.get("accion_sugerida"),
                item.get("modelo_usado"),
                item.get("created_at"),
            ]
            for item in reporte.get("validaciones_documentales", [])
        ],
    )

    _agregar_hoja_excel(
        workbook,
        "Métricas Finales",
        ["Indicador", "Valor", "Interpretación"],
        [
            [item.get("indicador"), item.get("valor"), item.get("interpretacion")]
            for item in metricas.get("indicadores_clave", [])
        ],
    )

    archivo = BytesIO()
    workbook.save(archivo)
    archivo.seek(0)
    return archivo


def _macroproceso_accion(accion: dict, evidencias_por_id: dict) -> str:
    origen_id = accion.get("origen_id")
    if accion.get("origen_tipo") == "macroproceso_evidencia" and origen_id in evidencias_por_id:
        return evidencias_por_id[origen_id].get("macroproceso") or "mejora_continua"
    if accion.get("origen_tipo") == "brecha_curricular" or accion.get("silabo_id") or accion.get("ciclo"):
        return "gestion_silabos"
    return accion.get("macroproceso") or "mejora_continua"


def _alerta_key(alerta: dict) -> tuple[str, str, str]:
    return (
        str(alerta.get("macroproceso") or ""),
        str(alerta.get("origen_tipo") or ""),
        str(alerta.get("origen_id") or alerta.get("codigo") or alerta.get("titulo") or ""),
    )


def _crear_alerta_candidata(
    macroproceso: str,
    origen_tipo: str,
    origen_id: Optional[str],
    codigo: Optional[str],
    titulo: str,
    descripcion: str,
    nivel_alerta: str,
    recomendacion: str,
) -> dict:
    return {
        "macroproceso": macroproceso,
        "origen_tipo": origen_tipo,
        "origen_id": origen_id,
        "codigo": codigo,
        "titulo": titulo,
        "descripcion": descripcion,
        "nivel_alerta": nivel_alerta,
        "estado": "activa",
        "recomendacion": recomendacion,
    }


def _ultimas_validaciones_por_evidencia(validaciones: list[dict]) -> dict:
    ordenadas = sorted(validaciones, key=lambda item: item.get("created_at") or "", reverse=True)
    ultimas = {}
    for item in ordenadas:
        evidencia_id = item.get("evidencia_id")
        if evidencia_id and evidencia_id not in ultimas:
            ultimas[evidencia_id] = item
    return ultimas


def _ultimos_analisis_por_macroproceso(analisis: list[dict]) -> dict:
    ordenados = sorted(analisis, key=lambda item: item.get("created_at") or "", reverse=True)
    ultimos = {}
    for item in ordenados:
        macroproceso = item.get("macroproceso")
        if macroproceso and macroproceso not in ultimos:
            ultimos[macroproceso] = item
    return ultimos


def _generar_candidatas_alertas() -> list[dict]:
    evidencias = _safe_select_table("macroproceso_evidencias")
    acciones = _safe_select_table("acciones_mejora")
    validaciones = _safe_select_table("validacion_ia_macroproceso_evidencias")
    brechas = _safe_select_table("brechas_curriculares")
    analisis = _safe_select_table("analisis_ia_macroprocesos")

    evidencias_por_id = {item.get("id"): item for item in evidencias if item.get("id")}
    ultimas_validaciones = _ultimas_validaciones_por_evidencia(validaciones)
    origenes_acciones = {
        str(item.get("origen_id"))
        for item in acciones
        if item.get("origen_id") and item.get("estado") not in {"atendida", "descartada"}
    }
    candidatos = []

    for evidencia in evidencias:
        evidencia_id = evidencia.get("id")
        macroproceso = evidencia.get("macroproceso") or "mejora_continua"
        codigo = evidencia.get("codigo")
        titulo = evidencia.get("titulo") or "Evidencia de macroproceso"
        prioridad = evidencia.get("prioridad")
        estado = evidencia.get("estado")
        avance = int(evidencia.get("avance") or 0)
        archivo_url = evidencia.get("archivo_url")

        if prioridad == "alta" and estado == "pendiente":
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "macroproceso_evidencia",
                evidencia_id,
                codigo,
                titulo,
                "Evidencia de prioridad alta pendiente de atencion.",
                "alta",
                "Asignar responsable y fecha de cierre para atender la evidencia.",
            ))
        if prioridad == "alta" and avance < 30:
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "macroproceso_evidencia_avance",
                evidencia_id,
                codigo,
                titulo,
                "Evidencia de prioridad alta con avance menor a 30%.",
                "alta",
                "Priorizar acciones inmediatas para incrementar el avance documentado.",
            ))
        if estado == "completado" and not archivo_url:
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "macroproceso_evidencia_sin_archivo",
                evidencia_id,
                codigo,
                titulo,
                "La evidencia figura como completada, pero no tiene archivo de sustento.",
                "critica",
                "Subir el archivo de sustento o corregir el estado de la evidencia.",
            ))
        if prioridad == "alta" and not archivo_url:
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "macroproceso_evidencia_sustento",
                evidencia_id,
                codigo,
                titulo,
                "Evidencia de prioridad alta sin archivo de sustento.",
                "alta",
                "Solicitar el documento de sustento al responsable.",
            ))
        if estado == "en_proceso" and avance < 50:
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "macroproceso_evidencia_en_proceso",
                evidencia_id,
                codigo,
                titulo,
                "Evidencia en proceso con avance menor a 50%.",
                "media",
                "Revisar el avance y registrar evidencias parciales verificables.",
            ))
        if estado == "observado":
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "macroproceso_evidencia_observada",
                evidencia_id,
                codigo,
                titulo,
                "La evidencia se encuentra observada y requiere subsanacion.",
                "critica",
                "Atender las observaciones y registrar el sustento corregido.",
            ))

        validacion = ultimas_validaciones.get(evidencia_id)
        if validacion:
            nivel = validacion.get("nivel_validez")
            pertinencia = validacion.get("pertinencia")
            if nivel == "bajo" or pertinencia == "no_corresponde":
                candidatos.append(_crear_alerta_candidata(
                    macroproceso,
                    "validacion_ia_evidencia",
                    evidencia_id,
                    codigo,
                    titulo,
                    "La ultima validacion IA detecto baja validez o falta de correspondencia documental.",
                    "critica",
                    validacion.get("accion_sugerida") or "Reemplazar o complementar el documento de sustento.",
                ))
            elif nivel == "medio":
                candidatos.append(_crear_alerta_candidata(
                    macroproceso,
                    "validacion_ia_evidencia",
                    evidencia_id,
                    codigo,
                    titulo,
                    "La ultima validacion IA indica validez documental parcial.",
                    "media",
                    validacion.get("accion_sugerida") or "Completar elementos documentales faltantes.",
                ))

    fecha_actual = datetime.now(timezone.utc).date()
    for accion in acciones:
        estado = accion.get("estado")
        prioridad = accion.get("prioridad")
        accion_id = accion.get("id")
        macroproceso = _macroproceso_accion(accion, evidencias_por_id)
        titulo = accion.get("titulo") or "Accion de mejora"
        fecha_limite = accion.get("fecha_limite")

        if prioridad == "alta" and estado == "pendiente":
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "accion_mejora",
                accion_id,
                None,
                titulo,
                "Accion de mejora de prioridad alta pendiente de atencion.",
                "alta",
                "Definir fecha de cierre y evidencias de cumplimiento.",
            ))
        if prioridad == "alta" and estado == "en_proceso":
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "accion_mejora_en_proceso",
                accion_id,
                None,
                titulo,
                "Accion de mejora de prioridad alta en proceso.",
                "media",
                "Monitorear el cierre y actualizar el avance de la accion.",
            ))
        if estado == "observada":
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "accion_mejora_observada",
                accion_id,
                None,
                titulo,
                "Accion de mejora observada.",
                "critica",
                "Subsanar la observacion de la accion de mejora.",
            ))
        if fecha_limite:
            try:
                fecha = datetime.fromisoformat(str(fecha_limite)).date()
                if fecha < fecha_actual and estado not in {"atendida", "descartada"}:
                    candidatos.append(_crear_alerta_candidata(
                        macroproceso,
                        "accion_mejora_vencida",
                        accion_id,
                        None,
                        titulo,
                        "Accion de mejora vencida segun la fecha limite registrada.",
                        "critica",
                        "Reprogramar o cerrar la accion con sustento verificable.",
                    ))
            except ValueError:
                pass

    for brecha in brechas:
        brecha_id = brecha.get("id")
        titulo = brecha.get("tipo_brecha") or "Brecha curricular"
        if brecha.get("prioridad") == "alta":
            candidatos.append(_crear_alerta_candidata(
                "gestion_silabos",
                "brecha_curricular",
                brecha_id,
                None,
                titulo,
                brecha.get("descripcion") or "Brecha curricular de prioridad alta.",
                "alta",
                brecha.get("recomendacion") or "Atender la brecha curricular con una accion de mejora.",
            ))
        if brecha_id and str(brecha_id) not in origenes_acciones:
            candidatos.append(_crear_alerta_candidata(
                "gestion_silabos",
                "brecha_curricular_sin_accion",
                brecha_id,
                None,
                titulo,
                "Brecha curricular sin accion de mejora asociada.",
                "media",
                "Generar o asociar una accion de mejora para cerrar la brecha.",
            ))

    ultimos_analisis = _ultimos_analisis_por_macroproceso(analisis)
    for macroproceso, item in ultimos_analisis.items():
        riesgo = item.get("nivel_riesgo")
        tipo_analisis = item.get("tipo_analisis") or "analisis_ia"
        if macroproceso == "mejora_continua" and riesgo == "alto":
            candidatos.append(_crear_alerta_candidata(
                "mejora_continua",
                "analisis_ia_integral",
                item.get("id"),
                None,
                "Riesgo alto en analisis integral de mejora continua",
                item.get("resumen") or "El ultimo analisis integral reporta riesgo alto.",
                "critica",
                "Revisar hallazgos integrales y priorizar acciones de mejora.",
            ))
        elif macroproceso in MACROPROCESOS_SEMAFORO and riesgo == "alto":
            candidatos.append(_crear_alerta_candidata(
                macroproceso,
                "analisis_ia_macroproceso",
                item.get("id"),
                None,
                f"Riesgo alto en {tipo_analisis}",
                item.get("resumen") or "El ultimo analisis IA del macroproceso reporta riesgo alto.",
                "alta",
                "Revisar el diagnostico IA y ejecutar acciones correctivas.",
            ))

    return candidatos


def _valor_historial(valor) -> Optional[str]:
    if valor is None:
        return None
    return str(valor)


def _registrar_historial_evidencia(
    evidencia: dict,
    campo: str,
    valor_anterior,
    valor_nuevo,
    observacion: Optional[str] = None,
    usuario: str = "backend",
) -> None:
    if _valor_historial(valor_anterior) == _valor_historial(valor_nuevo):
        return

    supabase.table("historial_macroproceso_evidencias").insert(
        {
            "evidencia_id": evidencia.get("id"),
            "macroproceso": evidencia.get("macroproceso"),
            "codigo": evidencia.get("codigo"),
            "titulo": evidencia.get("titulo"),
            "campo_modificado": campo,
            "valor_anterior": _valor_historial(valor_anterior),
            "valor_nuevo": _valor_historial(valor_nuevo),
            "observacion": observacion,
            "usuario": usuario,
        }
    ).execute()


def _registrar_historial_cambios(evidencia_actual: dict, data: dict, observacion: Optional[str] = None) -> None:
    for campo in CAMPOS_HISTORIAL:
        if campo not in data:
            continue
        _registrar_historial_evidencia(
            evidencia_actual,
            campo,
            evidencia_actual.get(campo),
            data.get(campo),
            observacion,
        )


def _texto_evidencia(evidencia: dict) -> str:
    partes = [
        evidencia.get("codigo"),
        evidencia.get("titulo"),
        evidencia.get("descripcion"),
        evidencia.get("tipo_evidencia"),
        evidencia.get("responsable"),
        evidencia.get("observacion"),
    ]
    return " ".join(str(parte or "") for parte in partes).lower()


def _es_evidencia_critica(evidencia: dict) -> bool:
    prioridad = evidencia.get("prioridad")
    estado = evidencia.get("estado")
    avance = int(evidencia.get("avance") or 0)
    texto = _texto_evidencia(evidencia)
    palabras_clave = ["acuerdo", "docente", "estudiante", "silabo", "sílabo", "portafolio", "plan anual"]

    return (
        (prioridad == "alta" and estado == "pendiente")
        or estado == "observado"
        or (avance < 30 and prioridad == "alta")
        or (estado == "completado" and not evidencia.get("archivo_url"))
        or (estado == "en_proceso" and avance < 50)
        or (estado == "pendiente" and any(palabra in texto for palabra in palabras_clave))
    )


def _descripcion_accion_evidencia(evidencia: dict) -> str:
    estado = evidencia.get("estado") or "sin estado"
    avance = evidencia.get("avance") or 0
    descripcion = (
        f"La evidencia {evidencia.get('codigo') or ''} presenta una condicion critica: "
        f"estado {estado}, prioridad {evidencia.get('prioridad') or 'media'} y avance {avance}%. "
        "Se requiere revisar, documentar y cerrar la evidencia dentro del ciclo de mejora continua."
    )
    if evidencia.get("estado") == "completado" and not evidencia.get("archivo_url"):
        descripcion += " La evidencia figura como completada, pero no tiene archivo de sustento."
    return descripcion


def _payload_accion_desde_evidencia(evidencia: dict) -> dict:
    prioridad = evidencia.get("prioridad") or "media"
    if evidencia.get("estado") == "observado":
        prioridad = "alta"
    if prioridad not in PRIORIDADES_VALIDAS:
        prioridad = "media"

    return {
        "origen_tipo": "macroproceso_evidencia",
        "origen_id": evidencia.get("id"),
        "titulo": f"Atender evidencia crítica: {evidencia.get('codigo') or '-'} - {evidencia.get('titulo') or 'Evidencia'}",
        "descripcion": _descripcion_accion_evidencia(evidencia),
        "recomendacion": "Asignar responsable, fecha de cierre y sustento documental verificable.",
        "prioridad": prioridad,
        "estado": "pendiente",
        "responsable": evidencia.get("responsable"),
        "evidencia_url": evidencia.get("archivo_url"),
        "observacion": f"Acción generada desde evidencia del macroproceso {evidencia.get('macroproceso')}.",
    }


def _accion_existente_para_evidencia(evidencia_id: str) -> list:
    response = (
        supabase.table("acciones_mejora")
        .select("*")
        .eq("origen_tipo", "macroproceso_evidencia")
        .eq("origen_id", evidencia_id)
        .execute()
    )
    return response.data or []


def _generar_accion_desde_evidencia(evidencia: dict) -> tuple[bool, dict | None]:
    existentes = _accion_existente_para_evidencia(evidencia.get("id"))
    if existentes:
        return False, existentes[0]

    payload = _payload_accion_desde_evidencia(evidencia)
    response = supabase.table("acciones_mejora").insert(payload).execute()
    accion = response.data[0] if response.data else payload
    return True, accion


def _guardar_historial_analisis_ia(
    macroproceso: str,
    tipo_analisis: str,
    resultado: dict,
    nivel_riesgo: Optional[str],
    resumen: Optional[str],
) -> None:
    try:
        supabase.table("analisis_ia_macroprocesos").insert(
            {
                "macroproceso": macroproceso,
                "tipo_analisis": tipo_analisis,
                "nivel_riesgo": nivel_riesgo,
                "resumen": resumen,
                "resultado_json": resultado,
                "modelo_usado": resultado.get("modelo_usado"),
                "usuario": "backend",
            }
        ).execute()
        print("[Historial IA] Análisis guardado correctamente para:", macroproceso)
    except Exception as e:
        print("[Historial IA] No se pudo guardar el análisis:", type(e).__name__, str(e))


def _resumen_historial_ia(item: dict) -> dict:
    return {
        "id": item.get("id"),
        "macroproceso": item.get("macroproceso"),
        "tipo_analisis": item.get("tipo_analisis"),
        "nivel_riesgo": item.get("nivel_riesgo"),
        "resumen": item.get("resumen"),
        "modelo_usado": item.get("modelo_usado"),
        "created_at": item.get("created_at"),
    }


def _comparar_riesgos(anterior: Optional[str], actual: Optional[str]) -> str:
    if anterior not in ORDEN_RIESGO or actual not in ORDEN_RIESGO:
        return "sin_datos"
    if ORDEN_RIESGO[actual] < ORDEN_RIESGO[anterior]:
        return "mejoro"
    if ORDEN_RIESGO[actual] > ORDEN_RIESGO[anterior]:
        return "empeoro"
    return "se_mantuvo"


@router.get("/evidencias")
def listar_evidencias(
    macroproceso: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    prioridad: Optional[str] = Query(None),
    responsable: Optional[str] = Query(None),
):
    try:
        query = supabase.table("macroproceso_evidencias").select("*")

        if macroproceso:
            query = query.eq("macroproceso", macroproceso.strip().lower())
        if estado:
            query = query.eq("estado", _normalizar_estado(estado))
        if prioridad:
            query = query.eq("prioridad", _normalizar_prioridad(prioridad))
        if responsable:
            query = query.ilike("responsable", f"%{responsable.strip()}%")

        response = query.execute()
        evidencias = _ordenar_evidencias(response.data or [])

        return {
            "message": "Evidencias de macroprocesos obtenidas correctamente",
            "data": evidencias,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/planificacion/analizar")
def analizar_planificacion_estrategica():
    try:
        resultado = ejecutar_grafo_planificacion()
        _guardar_historial_analisis_ia(
            macroproceso="planificacion_estrategica",
            tipo_analisis="analisis_planificacion",
            resultado=resultado,
            nivel_riesgo=resultado.get("nivel_riesgo"),
            resumen=resultado.get("resumen"),
        )

        return {
            "message": "An?lisis de planificaci?n estrat?gica generado correctamente",
            "data": resultado,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/gestion-academica/analizar")
def analizar_gestion_academica():
    try:
        resultado = ejecutar_grafo_gestion_academica()
        _guardar_historial_analisis_ia(
            macroproceso="gestion_academica",
            tipo_analisis="analisis_gestion_academica",
            resultado=resultado,
            nivel_riesgo=resultado.get("nivel_riesgo"),
            resumen=resultado.get("resumen"),
        )

        return {
            "message": "An?lisis de gesti?n acad?mica generado correctamente",
            "data": resultado,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/mejora-continua/analizar")
def analizar_mejora_continua():
    try:
        resultado = ejecutar_grafo_coordinador_mejora_continua()
        _guardar_historial_analisis_ia(
            macroproceso="mejora_continua",
            tipo_analisis="analisis_integral_mejora_continua",
            resultado=resultado,
            nivel_riesgo=resultado.get("nivel_riesgo_general"),
            resumen=resultado.get("resumen_general"),
        )

        return {
            "message": "An?lisis integral de mejora continua generado correctamente",
            "data": resultado,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/analisis-ia/historial")
def listar_historial_analisis_ia(
    macroproceso: Optional[str] = Query(None),
    tipo_analisis: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=100),
):
    try:
        query = supabase.table("analisis_ia_macroprocesos").select("*")

        if macroproceso:
            macroproceso_normalizado = macroproceso.strip().lower()
            if macroproceso_normalizado not in MACROPROCESOS_ANALISIS_VALIDOS:
                raise HTTPException(status_code=400, detail="Macroproceso de analisis IA no valido.")
            query = query.eq("macroproceso", macroproceso_normalizado)

        if tipo_analisis:
            tipo_normalizado = tipo_analisis.strip().lower()
            if tipo_normalizado not in TIPOS_ANALISIS_VALIDOS:
                raise HTTPException(status_code=400, detail="Tipo de analisis IA no valido.")
            query = query.eq("tipo_analisis", tipo_normalizado)

        response = query.order("created_at", desc=True).limit(limit).execute()
        return {
            "historial": [_resumen_historial_ia(item) for item in (response.data or [])],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/analisis-ia/historial/{analisis_id}")
def obtener_detalle_historial_analisis_ia(analisis_id: str):
    try:
        response = (
            supabase.table("analisis_ia_macroprocesos")
            .select("*")
            .eq("id", analisis_id)
            .execute()
        )
        if not response.data:
            raise HTTPException(status_code=404, detail="Analisis IA no encontrado.")

        return response.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/analisis-ia/comparar")
def comparar_historial_analisis_ia(macroproceso: str = Query(...)):
    try:
        macroproceso_normalizado = macroproceso.strip().lower()
        if macroproceso_normalizado not in MACROPROCESOS_ANALISIS_VALIDOS:
            raise HTTPException(status_code=400, detail="Macroproceso de analisis IA no valido.")

        response = (
            supabase.table("analisis_ia_macroprocesos")
            .select("*")
            .eq("macroproceso", macroproceso_normalizado)
            .order("created_at", desc=True)
            .limit(2)
            .execute()
        )
        registros = response.data or []
        actual = registros[0] if len(registros) >= 1 else None
        anterior = registros[1] if len(registros) >= 2 else None

        if not actual or not anterior:
            cambio = "sin_datos"
            resumen = "No hay dos analisis IA disponibles para comparar."
        else:
            cambio = _comparar_riesgos(anterior.get("nivel_riesgo"), actual.get("nivel_riesgo"))
            if cambio == "mejoro":
                resumen = f"El riesgo mejoro de {anterior.get('nivel_riesgo')} a {actual.get('nivel_riesgo')}."
            elif cambio == "empeoro":
                resumen = f"El riesgo empeoro de {anterior.get('nivel_riesgo')} a {actual.get('nivel_riesgo')}."
            elif cambio == "se_mantuvo":
                resumen = f"El riesgo se mantuvo en {actual.get('nivel_riesgo')}."
            else:
                resumen = "No se pudo determinar el cambio de riesgo."

        return {
            "macroproceso": macroproceso_normalizado,
            "analisis_actual": _resumen_historial_ia(actual) if actual else None,
            "analisis_anterior": _resumen_historial_ia(anterior) if anterior else None,
            "comparacion": {
                "riesgo_anterior": anterior.get("nivel_riesgo") if anterior else None,
                "riesgo_actual": actual.get("nivel_riesgo") if actual else None,
                "cambio_riesgo": cambio,
                "resumen": resumen,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/alertas/generar")
def generar_alertas_inteligentes():
    try:
        candidatas = _generar_candidatas_alertas()
        existentes_response = (
            supabase.table("alertas_inteligentes_macroprocesos")
            .select("*")
            .eq("estado", "activa")
            .execute()
        )
        existentes = existentes_response.data or []
        claves_existentes = {_alerta_key(item) for item in existentes}
        nuevas = []
        alertas_existentes = 0

        for alerta in candidatas:
            clave = _alerta_key(alerta)
            if clave in claves_existentes:
                alertas_existentes += 1
                continue
            nuevas.append(alerta)
            claves_existentes.add(clave)

        creadas = []
        if nuevas:
            insert_response = supabase.table("alertas_inteligentes_macroprocesos").insert(nuevas).execute()
            creadas = insert_response.data or nuevas

        alertas = sorted(
            creadas + existentes,
            key=lambda item: (
                ORDEN_ALERTA.get(item.get("nivel_alerta"), 99),
                item.get("created_at") or "",
            ),
        )

        return {
            "message": "Alertas inteligentes generadas correctamente",
            "alertas_creadas": len(creadas),
            "alertas_existentes": alertas_existentes,
            "alertas": alertas,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/alertas")
def listar_alertas_inteligentes(
    macroproceso: Optional[str] = Query(None),
    nivel_alerta: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
):
    try:
        query = supabase.table("alertas_inteligentes_macroprocesos").select("*")

        if macroproceso:
            query = query.eq("macroproceso", macroproceso.strip().lower())
        if nivel_alerta:
            query = query.eq("nivel_alerta", _normalizar_nivel_alerta(nivel_alerta))
        if estado:
            query = query.eq("estado", _normalizar_estado_alerta(estado))

        response = query.order("created_at", desc=True).limit(limit).execute()
        alertas = sorted(
            response.data or [],
            key=lambda item: (
                ORDEN_ALERTA.get(item.get("nivel_alerta"), 99),
                item.get("created_at") or "",
            ),
        )
        return {"alertas": alertas}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.put("/alertas/{alerta_id}")
def actualizar_alerta_inteligente(alerta_id: str, alerta: AlertaUpdate):
    try:
        data = alerta.model_dump(exclude_unset=True)
        if not data:
            raise HTTPException(status_code=400, detail="No hay datos para actualizar.")

        if "estado" in data:
            data["estado"] = _normalizar_estado_alerta(data.get("estado"))
        data["updated_at"] = _ahora_iso()

        response = (
            supabase.table("alertas_inteligentes_macroprocesos")
            .update(data)
            .eq("id", alerta_id)
            .execute()
        )
        if not response.data:
            raise HTTPException(status_code=404, detail="Alerta no encontrada.")

        return {
            "message": "Alerta actualizada correctamente",
            "data": response.data,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/semaforo")
def obtener_semaforo_macroprocesos():
    try:
        evidencias = _safe_select_table("macroproceso_evidencias")
        silabos = _safe_select_table("silabos")
        alertas_activas = (
            supabase.table("alertas_inteligentes_macroprocesos")
            .select("*")
            .eq("estado", "activa")
            .execute()
            .data
            or []
        )
        ultimos_analisis = _ultimos_analisis_por_macroproceso(_safe_select_table("analisis_ia_macroprocesos"))

        semaforos = []
        for macroproceso, nombre in MACROPROCESOS_SEMAFORO.items():
            if macroproceso == "gestion_silabos":
                total = len(silabos)
                completos = sum(1 for item in silabos if item.get("estado") == "completo")
                avance_promedio = round((completos / total) * 100) if total else 0
            else:
                evidencias_macro = [item for item in evidencias if item.get("macroproceso") == macroproceso]
                avance_promedio = _resumen_evidencias(evidencias_macro)["avance_promedio"]

            alertas_macro = [item for item in alertas_activas if item.get("macroproceso") == macroproceso]
            alertas_criticas = sum(1 for item in alertas_macro if item.get("nivel_alerta") == "critica")
            alertas_altas = sum(1 for item in alertas_macro if item.get("nivel_alerta") == "alta")
            riesgo_ia = (ultimos_analisis.get(macroproceso) or {}).get("nivel_riesgo") or "sin_datos"

            if alertas_criticas > 0 or avance_promedio < 40 or riesgo_ia == "alto":
                color = "rojo"
                mensaje = "Requiere atención prioritaria por bajo avance, alertas críticas o riesgo IA alto."
            elif avance_promedio >= 70 and alertas_criticas == 0 and alertas_altas <= 2:
                color = "verde"
                mensaje = "Cumplimiento aceptable, mantener seguimiento y evidencias actualizadas."
            else:
                color = "amarillo"
                mensaje = "Seguimiento preventivo requerido por avance intermedio o alertas activas."

            semaforos.append(
                {
                    "macroproceso": macroproceso,
                    "nombre": nombre,
                    "color": color,
                    "avance_promedio": avance_promedio,
                    "alertas_criticas": alertas_criticas,
                    "alertas_altas": alertas_altas,
                    "riesgo_ia": riesgo_ia,
                    "mensaje": mensaje,
                }
            )

        resumen = {
            "verde": sum(1 for item in semaforos if item["color"] == "verde"),
            "amarillo": sum(1 for item in semaforos if item["color"] == "amarillo"),
            "rojo": sum(1 for item in semaforos if item["color"] == "rojo"),
        }

        return {
            "semaforos": semaforos,
            "resumen": resumen,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/metricas-finales")
def obtener_metricas_finales_sistema():
    try:
        evidencias = _safe_select_table_reporte("macroproceso_evidencias")
        acciones = _safe_select_table_reporte("acciones_mejora")
        alertas = _safe_select_table_reporte("alertas_inteligentes_macroprocesos")
        analisis = _safe_select_table_reporte("analisis_ia_macroprocesos")
        analisis_silabos = _safe_select_table_reporte("analisis_silabo")
        validaciones = _safe_select_table_reporte("validacion_ia_macroproceso_evidencias")
        _safe_select_table_reporte("historial_macroproceso_evidencias")
        brechas = _safe_select_table_reporte("brechas_curriculares")
        silabos = _safe_select_table_reporte("silabos")
        seguimientos = _safe_select_table_reporte("seguimiento_semanal_evidencias")

        evidencias_por_id = {
            item.get("id"): item
            for item in evidencias
            if item.get("id")
        }
        alertas_activas = [item for item in alertas if item.get("estado") == "activa"]
        alertas_atendidas = [item for item in alertas if item.get("estado") == "atendida"]
        alertas_descartadas = [item for item in alertas if item.get("estado") == "descartada"]

        total_evidencias = len(evidencias)
        evidencias_con_archivo = sum(1 for item in evidencias if item.get("archivo_url"))
        evidencias_sin_archivo = total_evidencias - evidencias_con_archivo
        acciones_pendientes = sum(1 for item in acciones if item.get("estado") == "pendiente")
        acciones_en_proceso = sum(1 for item in acciones if item.get("estado") == "en_proceso")
        acciones_completadas = sum(
            1 for item in acciones if item.get("estado") in {"atendida", "completada", "completado"}
        )
        alertas_criticas = sum(1 for item in alertas_activas if item.get("nivel_alerta") == "critica")
        alertas_altas = sum(1 for item in alertas_activas if item.get("nivel_alerta") == "alta")
        alertas_medias = sum(1 for item in alertas_activas if item.get("nivel_alerta") == "media")
        alertas_bajas = sum(1 for item in alertas_activas if item.get("nivel_alerta") == "baja")
        validaciones_altas = sum(1 for item in validaciones if item.get("nivel_validez") == "alto")
        validaciones_medias = sum(1 for item in validaciones if item.get("nivel_validez") == "medio")
        validaciones_bajas = sum(1 for item in validaciones if item.get("nivel_validez") == "bajo")
        brechas_alta_prioridad = sum(1 for item in brechas if item.get("prioridad") == "alta")
        analisis_planificacion = sum(1 for item in analisis if item.get("tipo_analisis") == "analisis_planificacion")
        analisis_gestion_academica = sum(1 for item in analisis if item.get("tipo_analisis") == "analisis_gestion_academica")
        analisis_integral = sum(1 for item in analisis if item.get("tipo_analisis") == "analisis_integral_mejora_continua")
        acciones_automaticas = sum(1 for item in acciones if item.get("origen_tipo"))
        acciones_manuales = max(len(acciones) - acciones_automaticas, 0)
        total_analisis_ia = len(analisis) + len(analisis_silabos)
        fecha_limite_seguimiento = date.today() - timedelta(days=10)
        ultimos_seguimientos_por_evidencia = {}
        conteo_seguimientos_por_evidencia = {}
        for item in sorted(seguimientos, key=lambda seg: seg.get("semana_inicio") or "", reverse=True):
            evidencia_id = item.get("evidencia_id")
            if evidencia_id:
                conteo_seguimientos_por_evidencia[evidencia_id] = conteo_seguimientos_por_evidencia.get(evidencia_id, 0) + 1
            if evidencia_id and evidencia_id not in ultimos_seguimientos_por_evidencia:
                ultimos_seguimientos_por_evidencia[evidencia_id] = item
        evidencias_con_seguimiento = len(ultimos_seguimientos_por_evidencia)
        evidencias_sin_seguimiento = max(total_evidencias - evidencias_con_seguimiento, 0)
        seguimientos_requieren_apoyo = sum(1 for item in seguimientos if item.get("requiere_apoyo"))
        archivos_seguimiento_subidos = sum(1 for item in seguimientos if item.get("archivo_sustento_url"))
        evidencias_con_mas_de_un_seguimiento = sum(
            1 for total in conteo_seguimientos_por_evidencia.values() if total > 1
        )
        promedio_avance_semanal = round(
            sum(int(item.get("porcentaje_avance") or 0) for item in seguimientos) / len(seguimientos)
        ) if seguimientos else 0
        seguimientos_recientes = 0
        evidencias_estancadas = 0
        for item in ultimos_seguimientos_por_evidencia.values():
            try:
                fecha_inicio = date.fromisoformat(str(item.get("semana_inicio")))
                if fecha_inicio >= fecha_limite_seguimiento:
                    seguimientos_recientes += 1
            except ValueError:
                pass
            if item.get("nivel_avance") == "Sin avance" or int(item.get("porcentaje_avance") or 0) == 0:
                evidencias_estancadas += 1

        avance_general = round(
            sum(int(item.get("avance") or 0) for item in evidencias) / total_evidencias
        ) if total_evidencias else 0
        cobertura_documental = round((evidencias_con_archivo / total_evidencias) * 100) if total_evidencias else 0

        por_macroproceso = []
        for macroproceso, nombre in MACROPROCESOS_SEMAFORO.items():
            if macroproceso == "gestion_silabos":
                total_silabos = len(silabos)
                silabos_completos = sum(1 for item in silabos if item.get("estado") == "completo")
                avance_promedio = round((silabos_completos / total_silabos) * 100) if total_silabos else 0
                total_evidencias_macro = total_silabos
                evidencias_con_archivo_macro = sum(1 for item in silabos if item.get("archivo_url"))
                evidencias_sin_archivo_macro = total_silabos - evidencias_con_archivo_macro
            else:
                evidencias_macro = [
                    item for item in evidencias if item.get("macroproceso") == macroproceso
                ]
                total_evidencias_macro = len(evidencias_macro)
                avance_promedio = round(
                    sum(int(item.get("avance") or 0) for item in evidencias_macro) / total_evidencias_macro
                ) if total_evidencias_macro else 0
                evidencias_con_archivo_macro = sum(1 for item in evidencias_macro if item.get("archivo_url"))
                evidencias_sin_archivo_macro = total_evidencias_macro - evidencias_con_archivo_macro

            acciones_macro = [
                item
                for item in acciones
                if _macroproceso_accion(item, evidencias_por_id) == macroproceso
            ]
            alertas_macro = [
                item for item in alertas_activas if item.get("macroproceso") == macroproceso
            ]

            por_macroproceso.append(
                {
                    "macroproceso": macroproceso,
                    "nombre": nombre,
                    "total_evidencias": total_evidencias_macro,
                    "avance_promedio": avance_promedio,
                    "evidencias_con_archivo": evidencias_con_archivo_macro,
                    "evidencias_sin_archivo": evidencias_sin_archivo_macro,
                    "alertas_activas": len(alertas_macro),
                    "alertas_criticas": sum(1 for item in alertas_macro if item.get("nivel_alerta") == "critica"),
                    "acciones_mejora": len(acciones_macro),
                    "acciones_pendientes": sum(1 for item in acciones_macro if item.get("estado") == "pendiente"),
                    "analisis_ia": len(analisis_silabos)
                    if macroproceso == "gestion_silabos"
                    else sum(1 for item in analisis if item.get("macroproceso") == macroproceso),
                    "validaciones_ia": sum(
                        1 for item in validaciones if item.get("macroproceso") == macroproceso
                    ),
                }
            )

        indicadores_clave = [
            {
                "indicador": "Cobertura de evidencias con sustento",
                "valor": f"{cobertura_documental}%",
                "interpretacion": "Porcentaje de evidencias con archivo de sustento.",
            },
            {
                "indicador": "Cumplimiento promedio",
                "valor": f"{avance_general}%",
                "interpretacion": "Promedio general de avance de evidencias.",
            },
            {
                "indicador": "Acciones pendientes",
                "valor": acciones_pendientes,
                "interpretacion": "Cantidad de acciones de mejora que requieren seguimiento.",
            },
            {
                "indicador": "Alertas críticas",
                "valor": alertas_criticas,
                "interpretacion": "Cantidad de alertas que requieren atención prioritaria.",
            },
            {
                "indicador": "Uso de IA",
                "valor": total_analisis_ia + len(validaciones),
                "interpretacion": "Cantidad de análisis y validaciones generadas con IA.",
            },
        ]
        indicadores_clave.append(
            {
                "indicador": "Seguimientos semanales",
                "valor": len(seguimientos),
                "interpretacion": "Registros semanales de avance asociados a evidencias.",
            }
        )
        metricas_ia = {
            "analisis_ia_ejecutados": total_analisis_ia,
            "validaciones_documentales_ia": len(validaciones),
            "analisis_planificacion": analisis_planificacion,
            "analisis_gestion_academica": analisis_gestion_academica,
            "analisis_integral_mejora_continua": analisis_integral,
            "analisis_silabos": len(analisis_silabos),
            "validaciones_altas": validaciones_altas,
            "validaciones_medias": validaciones_medias,
            "validaciones_bajas": validaciones_bajas,
            "historial_ia_registrado": len(analisis),
        }
        metricas_acciones = {
            "total_acciones": len(acciones),
            "pendientes": acciones_pendientes,
            "en_proceso": acciones_en_proceso,
            "completadas": acciones_completadas,
            "acciones_automaticas": acciones_automaticas,
            "acciones_manuales": acciones_manuales,
        }
        metricas_alertas = {
            "total_activas": len(alertas_activas),
            "criticas": alertas_criticas,
            "altas": alertas_altas,
            "medias": alertas_medias,
            "bajas": alertas_bajas,
            "atendidas": len(alertas_atendidas),
            "descartadas": len(alertas_descartadas),
        }
        metricas_seguimiento = {
            "total_seguimientos": len(seguimientos),
            "evidencias_con_seguimiento": evidencias_con_seguimiento,
            "evidencias_sin_seguimiento": evidencias_sin_seguimiento,
            "seguimientos_recientes": seguimientos_recientes,
            "requieren_apoyo": seguimientos_requieren_apoyo,
            "promedio_avance_semanal": promedio_avance_semanal,
            "evidencias_estancadas": evidencias_estancadas,
            "archivos_seguimiento_subidos": archivos_seguimiento_subidos,
            "evidencias_con_mas_de_un_seguimiento": evidencias_con_mas_de_un_seguimiento,
        }

        return {
            "message": "Métricas finales generadas correctamente",
            "data": {
                "resumen_general": {
                    "total_macroprocesos": len(MACROPROCESOS_SEMAFORO),
                    "total_evidencias": total_evidencias,
                    "evidencias_con_archivo": evidencias_con_archivo,
                    "evidencias_sin_archivo": evidencias_sin_archivo,
                    "total_acciones_mejora": len(acciones),
                    "acciones_pendientes": acciones_pendientes,
                    "acciones_en_proceso": acciones_en_proceso,
                    "acciones_completadas": acciones_completadas,
                    "total_alertas_activas": len(alertas_activas),
                    "alertas_criticas": alertas_criticas,
                    "total_analisis_ia": total_analisis_ia,
                    "total_validaciones_ia": len(validaciones),
                    "validaciones_altas": validaciones_altas,
                    "validaciones_medias": validaciones_medias,
                    "validaciones_bajas": validaciones_bajas,
                    "total_brechas": len(brechas),
                    "brechas_alta_prioridad": brechas_alta_prioridad,
                    "total_silabos": len(silabos),
                    "total_seguimientos_semanales": len(seguimientos),
                    "total_archivos_seguimiento": archivos_seguimiento_subidos,
                    "seguimientos_requieren_apoyo": seguimientos_requieren_apoyo,
                },
                "por_macroproceso": por_macroproceso,
                "metricas_ia": metricas_ia,
                "metricas_acciones": metricas_acciones,
                "metricas_alertas": metricas_alertas,
                "metricas_seguimiento": metricas_seguimiento,
                "indicadores_clave": indicadores_clave,
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/encuesta-aceptacion")
def registrar_encuesta_aceptacion(encuesta: EncuestaAceptacionCreate):
    try:
        data = encuesta.model_dump()
        response = supabase.table("encuesta_aceptacion").insert(data).execute()

        return {
            "message": "Encuesta registrada correctamente",
            "data": response.data,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/encuesta-aceptacion/resultados")
def obtener_resultados_encuesta_aceptacion():
    try:
        respuestas = _safe_select_table_reporte("encuesta_aceptacion")
        campos = [
            "claridad_interfaz",
            "facilidad_uso",
            "utilidad_evidencias",
            "utilidad_ia",
            "utilidad_alertas",
            "utilidad_semaforo",
            "satisfaccion_general",
        ]
        total = len(respuestas)
        promedios = {}

        for campo in campos:
            promedios[campo] = round(
                sum(int(item.get(campo) or 0) for item in respuestas) / total,
                2,
            ) if total else 0

        promedio_general = round(
            sum(promedios.values()) / len(campos),
            2,
        ) if total else 0

        comentarios = [
            {
                "nombre_usuario": item.get("nombre_usuario"),
                "rol_usuario": item.get("rol_usuario"),
                "comentario": item.get("comentario"),
                "created_at": item.get("created_at"),
            }
            for item in sorted(respuestas, key=lambda row: row.get("created_at") or "", reverse=True)
            if item.get("comentario")
        ]

        return {
            "total_respuestas": total,
            "promedios": promedios,
            "promedio_general": promedio_general,
            "comentarios": comentarios,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/reporte-integral")
def obtener_reporte_integral_mejora_continua():
    try:
        print("[Reporte Integral] Generando reporte...")

        semaforo_data = obtener_semaforo_macroprocesos()
        semaforo = semaforo_data.get("semaforos", [])

        evidencias = _safe_select_table_reporte("macroproceso_evidencias")
        acciones = _safe_select_table_reporte("acciones_mejora")
        analisis = _safe_select_table_reporte("analisis_ia_macroprocesos")
        validaciones = _safe_select_table_reporte("validacion_ia_macroproceso_evidencias")
        brechas = _safe_select_table_reporte("brechas_curriculares")
        seguimientos = _safe_select_table_reporte("seguimiento_semanal_evidencias")

        alertas_activas = [
            item
            for item in _safe_select_table_reporte("alertas_inteligentes_macroprocesos")
            if item.get("estado") == "activa"
        ]

        ultimos_analisis_map = _ultimos_analisis_por_macroproceso(analisis)
        ultimos_analisis = [
            item
            for macroproceso, item in ultimos_analisis_map.items()
            if macroproceso in MACROPROCESOS_ANALISIS_VALIDOS
        ]

        ultimas_validaciones_map = _ultimas_validaciones_por_evidencia(validaciones)
        validaciones_documentales = list(ultimas_validaciones_map.values())

        evidencias_criticas = [
            item
            for item in evidencias
            if item.get("estado") in {"pendiente", "observado"}
            or item.get("prioridad") == "alta"
            or int(item.get("avance") or 0) < 50
            or not item.get("archivo_url")
        ]

        total_alertas_criticas = sum(
            1 for item in alertas_activas if item.get("nivel_alerta") == "critica"
        )
        acciones_pendientes = sum(1 for item in acciones if item.get("estado") == "pendiente")
        acciones_en_proceso = sum(1 for item in acciones if item.get("estado") == "en_proceso")
        acciones_completadas = sum(
            1 for item in acciones if item.get("estado") in {"atendida", "completada"}
        )
        evidencias_por_id_reporte = {item.get("id"): item for item in evidencias if item.get("id")}
        seguimientos_enriquecidos = []
        for item in seguimientos:
            evidencia = evidencias_por_id_reporte.get(item.get("evidencia_id"), {})
            seguimientos_enriquecidos.append({
                **item,
                "titulo_evidencia": evidencia.get("titulo"),
                "estado_evidencia": evidencia.get("estado"),
                "prioridad_evidencia": evidencia.get("prioridad"),
            })

        recomendaciones = []
        if total_alertas_criticas:
            recomendaciones.append("Atender de manera prioritaria las alertas críticas activas.")
        if any(not item.get("archivo_url") for item in evidencias):
            recomendaciones.append("Completar la carga documental de evidencias pendientes.")
        if acciones_pendientes:
            recomendaciones.append("Dar seguimiento a las acciones de mejora pendientes.")
        if any(item.get("color") == "rojo" for item in semaforo):
            recomendaciones.append("Priorizar la revisión del macroproceso con semáforo rojo.")
        if any(item.get("nivel_validez") == "bajo" for item in validaciones_documentales):
            recomendaciones.append("Revisar la calidad documental de las evidencias con baja validez.")
        if not recomendaciones:
            recomendaciones.append("Mantener el seguimiento periódico de evidencias, alertas y acciones de mejora.")

        reporte = {
            "fecha_generacion": _ahora_iso(),
            "titulo": "Reporte Integral de Mejora Continua",
            "resumen": {
                "total_macroprocesos": len(MACROPROCESOS_SEMAFORO),
                "total_evidencias": len(evidencias),
                "total_alertas_activas": len(alertas_activas),
                "total_alertas_criticas": total_alertas_criticas,
                "total_acciones_mejora": len(acciones),
                "acciones_pendientes": acciones_pendientes,
                "acciones_en_proceso": acciones_en_proceso,
                "acciones_completadas": acciones_completadas,
                "total_analisis_ia": len(ultimos_analisis),
                "total_validaciones_ia": len(validaciones_documentales),
                "total_seguimientos_semanales": len(seguimientos),
                "total_archivos_seguimiento": sum(1 for item in seguimientos if item.get("archivo_sustento_url")),
            },
            "semaforo": semaforo,
            "evidencias": evidencias,
            "evidencias_criticas": evidencias_criticas,
            "alertas_activas": sorted(
                alertas_activas,
                key=lambda item: (
                    ORDEN_ALERTA.get(item.get("nivel_alerta"), 99),
                    item.get("created_at") or "",
                ),
            ),
            "acciones_mejora": acciones,
            "ultimos_analisis_ia": ultimos_analisis,
            "validaciones_documentales": validaciones_documentales,
            "brechas_curriculares": brechas,
            "seguimientos_semanales": seguimientos_enriquecidos,
            "recomendaciones_generales": recomendaciones,
        }

        print("[Reporte Integral] Reporte generado correctamente.")
        return reporte
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/reporte-integral/excel")
def exportar_reporte_integral_excel():
    try:
        print("[Reporte Integral Excel] Generando archivo Excel...")
        reporte = obtener_reporte_integral_mejora_continua()
        metricas = obtener_metricas_finales_sistema()
        archivo = _crear_excel_reporte_integral(reporte, metricas)

        headers = {
            "Content-Disposition": 'attachment; filename="reporte_integral_mejora_continua.xlsx"'
        }
        return StreamingResponse(
            archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/evidencias/{evidencia_id}")
def obtener_evidencia(evidencia_id: str):
    try:
        evidencia = _obtener_evidencia_o_404(evidencia_id)

        return {
            "message": "Evidencia obtenida correctamente",
            "data": evidencia,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/evidencias/{evidencia_id}/seguimiento-semanal")
def crear_seguimiento_semanal(evidencia_id: str, seguimiento: SeguimientoSemanalCreate):
    try:
        evidencia = _obtener_evidencia_o_404(evidencia_id)
        data = seguimiento.model_dump()
        data["evidencia_id"] = evidencia_id
        data["macroproceso"] = evidencia.get("macroproceso")
        data["codigo_evidencia"] = evidencia.get("codigo")
        data["updated_at"] = _ahora_iso()
        _validar_payload_seguimiento(data)

        existente = (
            supabase.table("seguimiento_semanal_evidencias")
            .select("id")
            .eq("evidencia_id", evidencia_id)
            .eq("semana_inicio", data["semana_inicio"])
            .execute()
        )
        if existente.data:
            raise HTTPException(
                status_code=400,
                detail="Ya existe un seguimiento semanal para esta evidencia en la semana indicada.",
            )

        response = supabase.table("seguimiento_semanal_evidencias").insert(data).execute()
        seguimiento_creado = response.data[0] if response.data else data

        supabase.table("macroproceso_evidencias").update({"updated_at": _ahora_iso()}).eq("id", evidencia_id).execute()
        try:
            _registrar_historial_evidencia(
                evidencia,
                "seguimiento_semanal",
                None,
                data.get("nivel_avance"),
                observacion=f"Seguimiento semanal registrado: {data.get('porcentaje_avance', 0)}% de avance.",
            )
        except Exception as historial_error:
            print("[Macroprocesos] No se pudo registrar historial de seguimiento:", type(historial_error).__name__, str(historial_error))

        return {
            "message": "Seguimiento semanal registrado correctamente",
            "data": seguimiento_creado,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/evidencias/{evidencia_id}/seguimiento-semanal")
def listar_seguimiento_semanal(evidencia_id: str):
    try:
        _obtener_evidencia_o_404(evidencia_id)
        response = (
            supabase.table("seguimiento_semanal_evidencias")
            .select("*")
            .eq("evidencia_id", evidencia_id)
            .order("semana_inicio", desc=True)
            .execute()
        )
        return {
            "message": "Seguimientos semanales obtenidos correctamente",
            "data": response.data or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/evidencias/{evidencia_id}/seguimiento-semanal/ultimo")
def obtener_ultimo_seguimiento_semanal(evidencia_id: str):
    try:
        _obtener_evidencia_o_404(evidencia_id)
        response = (
            supabase.table("seguimiento_semanal_evidencias")
            .select("*")
            .eq("evidencia_id", evidencia_id)
            .order("semana_inicio", desc=True)
            .limit(1)
            .execute()
        )
        ultimo = response.data[0] if response.data else None
        return {
            "message": "Último seguimiento semanal obtenido correctamente",
            "data": ultimo,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.put("/evidencias/seguimiento-semanal/{seguimiento_id}")
def actualizar_seguimiento_semanal(seguimiento_id: str, seguimiento: SeguimientoSemanalUpdate):
    try:
        seguimiento_actual = _obtener_seguimiento_o_404(seguimiento_id)
        data = seguimiento.model_dump(exclude_unset=True)
        if not data:
            raise HTTPException(status_code=400, detail="No hay datos para actualizar.")

        payload_validacion = {**seguimiento_actual, **data}
        _validar_payload_seguimiento(payload_validacion)
        data["updated_at"] = _ahora_iso()

        if "semana_inicio" in data:
            duplicado = (
                supabase.table("seguimiento_semanal_evidencias")
                .select("id")
                .eq("evidencia_id", seguimiento_actual.get("evidencia_id"))
                .eq("semana_inicio", data["semana_inicio"])
                .neq("id", seguimiento_id)
                .execute()
            )
            if duplicado.data:
                raise HTTPException(
                    status_code=400,
                    detail="Ya existe otro seguimiento semanal para esta evidencia en la semana indicada.",
                )

        response = (
            supabase.table("seguimiento_semanal_evidencias")
            .update(data)
            .eq("id", seguimiento_id)
            .execute()
        )
        return {
            "message": "Seguimiento semanal actualizado correctamente",
            "data": response.data[0] if response.data else {**seguimiento_actual, **data},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.delete("/evidencias/seguimiento-semanal/{seguimiento_id}")
def eliminar_seguimiento_semanal(seguimiento_id: str):
    try:
        _obtener_seguimiento_o_404(seguimiento_id)
        supabase.table("seguimiento_semanal_evidencias").delete().eq("id", seguimiento_id).execute()
        return {
            "message": "Seguimiento semanal eliminado correctamente",
            "data": {"id": seguimiento_id},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/evidencias/seguimiento-semanal/{seguimiento_id}/archivo")
async def subir_archivo_seguimiento_semanal(seguimiento_id: str, archivo: UploadFile = File(...)):
    try:
        seguimiento = _obtener_seguimiento_o_404(seguimiento_id)
        evidencia = _obtener_evidencia_o_404(seguimiento.get("evidencia_id"))
        nombre_original = archivo.filename or ""
        extension = os.path.splitext(nombre_original)[1].lower()
        content_type = (archivo.content_type or "").lower()

        extension_valida = extension in EXTENSIONES_EVIDENCIA
        content_type_valido = content_type in TIPOS_CONTENIDO_EVIDENCIA
        if not extension_valida and not (not extension and content_type_valido):
            raise HTTPException(
                status_code=400,
                detail="Formato no permitido. Solo se aceptan archivos PDF, DOC, DOCX, XLSX, PNG, JPG o JPEG.",
            )

        contenido = await archivo.read()
        if not contenido:
            raise HTTPException(status_code=400, detail="El archivo está vacío.")

        nombre_seguro = os.path.basename(nombre_original).replace(" ", "_") or f"seguimiento{extension}"
        codigo = (evidencia.get("codigo") or seguimiento.get("evidencia_id") or seguimiento_id).replace("/", "-").replace("\\", "-")
        ruta_storage = (
            f"macroprocesos/{evidencia.get('macroproceso')}/{codigo}/seguimientos/"
            f"{uuid.uuid4()}_{nombre_seguro}"
        )

        try:
            supabase.storage.from_(BUCKET_EVIDENCIAS).upload(
                path=ruta_storage,
                file=contenido,
                file_options={
                    "content-type": archivo.content_type or "application/octet-stream",
                    "upsert": "true",
                },
            )
        except Exception as storage_error:
            raise HTTPException(
                status_code=500,
                detail=(
                    "No se pudo subir el archivo de seguimiento. "
                    f"Verifique que el bucket '{BUCKET_EVIDENCIAS}' exista y tenga permisos de Storage."
                ),
            ) from storage_error

        archivo_url = supabase.storage.from_(BUCKET_EVIDENCIAS).get_public_url(ruta_storage)
        response = (
            supabase.table("seguimiento_semanal_evidencias")
            .update({
                "archivo_sustento_url": archivo_url,
                "updated_at": _ahora_iso(),
            })
            .eq("id", seguimiento_id)
            .execute()
        )

        return {
            "message": "Archivo de seguimiento subido correctamente",
            "data": response.data[0] if response.data else {**seguimiento, "archivo_sustento_url": archivo_url},
            "archivo_url": archivo_url,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="No se pudo subir el archivo de seguimiento.") from e


@router.post("/evidencias/{evidencia_id}/validar-ia")
def validar_evidencia_con_ia(evidencia_id: str):
    try:
        evidencia = _obtener_evidencia_o_404(evidencia_id)
        if not evidencia.get("archivo_url"):
            raise HTTPException(
                status_code=400,
                detail="Primero suba un archivo de sustento para validar esta evidencia.",
            )

        try:
            resultado = ejecutar_grafo_validacion_evidencia(evidencia_id)
        except ValueError as error:
            raise HTTPException(status_code=400, detail=str(error)) from error

        return {
            "message": "Validacion IA de evidencia generada correctamente",
            "data": resultado,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/evidencias/{evidencia_id}/validacion-ia")
def obtener_ultima_validacion_evidencia_ia(evidencia_id: str):
    try:
        _obtener_evidencia_o_404(evidencia_id)
        response = (
            supabase.table("validacion_ia_macroproceso_evidencias")
            .select("*")
            .eq("evidencia_id", evidencia_id)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
        )

        return {
            "message": "Ultima validacion IA obtenida correctamente"
            if response.data
            else "No hay validaciones IA registradas para esta evidencia.",
            "data": response.data[0] if response.data else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/evidencias/{evidencia_id}/validacion-ia/historial")
def obtener_historial_validacion_evidencia_ia(evidencia_id: str):
    try:
        _obtener_evidencia_o_404(evidencia_id)
        response = (
            supabase.table("validacion_ia_macroproceso_evidencias")
            .select("*")
            .eq("evidencia_id", evidencia_id)
            .order("created_at", desc=True)
            .execute()
        )

        return {
            "evidencia_id": evidencia_id,
            "historial": response.data or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/evidencias")
def crear_evidencia(evidencia: EvidenciaCreate):
    try:
        data = _normalizar_payload(evidencia.model_dump())
        macroproceso = data.get("macroproceso")
        codigo = (data.get("codigo") or "").strip()

        if macroproceso not in MACROPROCESOS_EVIDENCIAS_VALIDOS:
            raise HTTPException(
                status_code=400,
                detail="El macroproceso debe ser planificacion_estrategica o gestion_academica.",
            )

        if not codigo:
            raise HTTPException(status_code=400, detail="El código de evidencia es obligatorio.")

        data["codigo"] = codigo

        existente = (
            supabase.table("macroproceso_evidencias")
            .select("id")
            .eq("macroproceso", macroproceso)
            .eq("codigo", codigo)
            .limit(1)
            .execute()
        )
        if existente.data:
            raise HTTPException(
                status_code=400,
                detail="Ya existe una evidencia con ese código en este macroproceso.",
            )

        response = supabase.table("macroproceso_evidencias").insert(data).execute()
        evidencia_creada = response.data[0] if response.data else data

        try:
            _registrar_historial_evidencia(
                evidencia_creada,
                "creacion",
                None,
                "Evidencia registrada",
                "Registro inicial de evidencia manual",
            )
        except Exception as historial_error:
            print("[Macroprocesos] No se pudo guardar historial inicial:", type(historial_error).__name__)

        print(f"[Macroprocesos] Evidencia registrada correctamente: {codigo}")

        return {
            "message": "Evidencia registrada correctamente",
            "data": response.data,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.put("/evidencias/{evidencia_id}")
def actualizar_evidencia(evidencia_id: str, evidencia: EvidenciaUpdate):
    try:
        evidencia_actual = _obtener_evidencia_o_404(evidencia_id)
        data = evidencia.model_dump(exclude_unset=True)
        data = _normalizar_payload(data)

        if not data:
            raise HTTPException(status_code=400, detail="No hay datos para actualizar.")

        data["updated_at"] = _ahora_iso()

        response = (
            supabase.table("macroproceso_evidencias")
            .update(data)
            .eq("id", evidencia_id)
            .execute()
        )

        if not response.data:
            raise HTTPException(status_code=404, detail="Evidencia no encontrada.")

        _registrar_historial_cambios(
            evidencia_actual,
            data,
            observacion="Actualización de evidencia desde el sistema",
        )

        return {
            "message": "Evidencia actualizada correctamente",
            "data": response.data,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/evidencias/{evidencia_id}/historial")
def obtener_historial_evidencia(evidencia_id: str):
    try:
        _obtener_evidencia_o_404(evidencia_id)
        response = (
            supabase.table("historial_macroproceso_evidencias")
            .select("*")
            .eq("evidencia_id", evidencia_id)
            .order("created_at", desc=True)
            .execute()
        )

        return {
            "evidencia_id": evidencia_id,
            "historial": response.data or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/evidencias/{evidencia_id}/archivo")
async def subir_archivo_evidencia(evidencia_id: str, archivo: UploadFile = File(...)):
    try:
        evidencia = _obtener_evidencia_o_404(evidencia_id)
        nombre_original = archivo.filename or ""
        extension = os.path.splitext(nombre_original)[1].lower()
        content_type = (archivo.content_type or "").lower()

        extension_valida = extension in EXTENSIONES_EVIDENCIA
        content_type_valido = content_type in TIPOS_CONTENIDO_EVIDENCIA
        if not extension_valida and not (not extension and content_type_valido):
            raise HTTPException(
                status_code=400,
                detail="Formato no permitido. Solo se aceptan archivos PDF, DOC, DOCX, XLSX, PNG, JPG o JPEG.",
            )

        contenido = await archivo.read()
        if not contenido:
            raise HTTPException(status_code=400, detail="El archivo está vacío.")

        nombre_seguro = os.path.basename(nombre_original).replace(" ", "_") or f"evidencia{extension}"
        codigo = (evidencia.get("codigo") or evidencia_id).replace("/", "-").replace("\\", "-")
        ruta_storage = (
            f"macroprocesos/{evidencia.get('macroproceso')}/{codigo}/"
            f"{uuid.uuid4()}_{nombre_seguro}"
        )

        try:
            supabase.storage.from_(BUCKET_EVIDENCIAS).upload(
                path=ruta_storage,
                file=contenido,
                file_options={
                    "content-type": archivo.content_type or "application/octet-stream",
                    "upsert": "true",
                },
            )
        except Exception as storage_error:
            raise HTTPException(
                status_code=500,
                detail=(
                    "No se pudo subir el archivo de evidencia. "
                    f"Verifique que el bucket '{BUCKET_EVIDENCIAS}' exista y que la clave de Supabase tenga permisos de Storage."
                ),
            ) from storage_error

        archivo_url = supabase.storage.from_(BUCKET_EVIDENCIAS).get_public_url(ruta_storage)
        response = (
            supabase.table("macroproceso_evidencias")
            .update({
                "archivo_url": archivo_url,
                "updated_at": _ahora_iso(),
            })
            .eq("id", evidencia_id)
            .execute()
        )

        try:
            _registrar_historial_evidencia(
                evidencia,
                "archivo_url",
                evidencia.get("archivo_url"),
                archivo_url,
                observacion="Archivo de sustento subido",
            )
        except Exception as historial_error:
            print("[Macroprocesos] No se pudo registrar historial de archivo:", type(historial_error).__name__, str(historial_error))

        evidencia_actualizada = response.data[0] if response.data else {}
        return {
            "message": "Archivo de evidencia subido correctamente",
            "data": {
                "id": evidencia_id,
                "archivo_url": archivo_url,
                "nombre_archivo": nombre_original,
                **evidencia_actualizada,
            },
            "archivo_url": archivo_url,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="No se pudo subir el archivo de evidencia.") from e


@router.post("/evidencias/{evidencia_id}/generar-accion")
def generar_accion_desde_evidencia(evidencia_id: str):
    try:
        evidencia = _obtener_evidencia_o_404(evidencia_id)
        creada, accion = _generar_accion_desde_evidencia(evidencia)

        if not creada:
            return {
                "message": "Ya existe una acción de mejora para esta evidencia.",
                "accion_existente": accion,
                "data": accion,
            }

        return {
            "message": "Acción de mejora generada correctamente.",
            "data": accion,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/{macroproceso}/generar-acciones-desde-evidencias")
def generar_acciones_desde_evidencias_macroproceso(macroproceso: str):
    try:
        macroproceso_normalizado = macroproceso.strip().lower()
        if macroproceso_normalizado not in MACROPROCESOS_ACCIONES_VALIDOS:
            raise HTTPException(
                status_code=400,
                detail="Macroproceso no permitido. Use planificacion_estrategica o gestion_academica.",
            )

        response = (
            supabase.table("macroproceso_evidencias")
            .select("*")
            .eq("macroproceso", macroproceso_normalizado)
            .execute()
        )
        evidencias = response.data or []
        criticas = [evidencia for evidencia in evidencias if _es_evidencia_critica(evidencia)]
        acciones = []
        acciones_creadas = 0
        acciones_existentes = 0

        for evidencia in criticas:
            creada, accion = _generar_accion_desde_evidencia(evidencia)
            if creada:
                acciones_creadas += 1
            else:
                acciones_existentes += 1
            if accion:
                acciones.append(accion)

        return {
            "message": "Acciones de mejora generadas correctamente",
            "acciones_creadas": acciones_creadas,
            "acciones_existentes": acciones_existentes,
            "acciones": acciones,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.delete("/evidencias/{evidencia_id}")
def eliminar_evidencia(evidencia_id: str):
    try:
        response = (
            supabase.table("macroproceso_evidencias")
            .delete()
            .eq("id", evidencia_id)
            .execute()
        )

        if not response.data:
            raise HTTPException(status_code=404, detail="Evidencia no encontrada.")

        return {
            "message": "Evidencia eliminada correctamente",
            "data": response.data,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/dashboard")
def obtener_dashboard_macroprocesos():
    try:
        response = supabase.table("macroproceso_evidencias").select("*").execute()
        evidencias = response.data or []
        resumen = _resumen_evidencias(evidencias)

        return {
            "message": "Dashboard de macroprocesos obtenido correctamente",
            "data": {
                "total_evidencias": resumen["total"],
                "total_planificacion_estrategica": sum(
                    1 for item in evidencias if item.get("macroproceso") == "planificacion_estrategica"
                ),
                "total_gestion_academica": sum(
                    1 for item in evidencias if item.get("macroproceso") == "gestion_academica"
                ),
                "pendientes": resumen["pendientes"],
                "en_proceso": resumen["en_proceso"],
                "completadas": resumen["completadas"],
                "observadas": resumen["observadas"],
                "avance_promedio": resumen["avance_promedio"],
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/dashboard/{macroproceso}")
def obtener_dashboard_macroproceso(macroproceso: str):
    try:
        macroproceso_normalizado = macroproceso.strip().lower()
        response = (
            supabase.table("macroproceso_evidencias")
            .select("*")
            .eq("macroproceso", macroproceso_normalizado)
            .execute()
        )
        resumen = _resumen_evidencias(response.data or [])

        return {
            "message": "Dashboard del macroproceso obtenido correctamente",
            "data": resumen,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e
