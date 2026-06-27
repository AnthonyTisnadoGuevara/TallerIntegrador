from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


COLOR_DARK = "1E3A8A"
COLOR_BLUE = "DBEAFE"
COLOR_GREEN = "DCFCE7"
COLOR_RED = "FEE2E2"
COLOR_YELLOW = "FEF3C7"
COLOR_PURPLE = "EDE9FE"
COLOR_GRAY = "F8FAFC"
COLOR_BORDER = "CBD5E1"
FOOTER_TEXT = "Reporte generado automáticamente por el Sistema de Monitoreo del Plan de Mejora Continua - ISIA."


def valor_excel(valor: Any) -> Any:
    if isinstance(valor, (dict, list)):
        return str(valor)
    if valor is None:
        return ""
    return valor


def aplicar_estilo_titulo(ws, rango: str, texto: str) -> None:
    ws.merge_cells(rango)
    celda = ws[rango.split(":")[0]]
    celda.value = texto
    celda.fill = PatternFill("solid", fgColor=COLOR_DARK)
    celda.font = Font(bold=True, color="FFFFFF", size=16)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90


def preparar_hoja(ws, orientacion_horizontal: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    if orientacion_horizontal:
        ws.page_setup.orientation = "landscape"


def crear_titulo_reporte(ws, titulo: str, fecha: str) -> None:
    preparar_hoja(ws)
    aplicar_estilo_titulo(ws, "A1:J1", titulo)
    ws["A2"] = "Sistema de Monitoreo del Plan de Mejora Continua - ISIA"
    ws["A3"] = "Programa de Ingeniería de Sistemas e Inteligencia Artificial"
    ws["A4"] = f"Fecha de generación: {fecha or ''}"
    ws["H4"] = "Versión: v1.0"
    for ref in ("A2", "A3", "A4", "H4"):
        ws[ref].font = Font(bold=ref in {"A2", "A3"}, color="334155")
        ws[ref].alignment = Alignment(vertical="center")


def agregar_interpretacion(ws, row: int, textos: list[str]) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row, 1, "Interpretación ejecutiva")
    ws.cell(row, 1).font = Font(bold=True, color="1E3A8A", size=12)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLOR_BLUE)
    contenido = "\n".join(f"• {texto}" for texto in textos) if textos else "• No se identifican observaciones críticas en los indicadores consolidados."
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 4, end_column=10)
    celda = ws.cell(row + 1, 1, contenido)
    celda.alignment = Alignment(wrap_text=True, vertical="top")
    celda.fill = PatternFill("solid", fgColor=COLOR_GRAY)


def aplicar_estilo_tabla(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    preparar_hoja(ws, orientacion_horizontal=max_col > 8)
    border = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[min_row]:
        cell.fill = PatternFill("solid", fgColor=COLOR_DARK)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ajustar_ancho_columnas(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 14), 58)


def colorear_estado(cell) -> None:
    estado = str(cell.value or "").lower()
    if estado in {"completo", "completado", "atendida", "verde"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
    elif estado in {"pendiente", "incompleto", "rojo", "activa"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_RED)
    elif estado in {"observado", "en_proceso", "amarillo"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)


def colorear_prioridad(cell) -> None:
    prioridad = str(cell.value or "").lower()
    if prioridad in {"alta", "critica", "crítica"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_RED)
    elif prioridad == "media":
        cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)
    elif prioridad == "baja":
        cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)


def colorear_riesgo(cell) -> None:
    riesgo = str(cell.value or "").lower()
    if riesgo in {"alto", "rojo", "critica", "crítica"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_RED)
    elif riesgo in {"medio", "amarillo"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)
    elif riesgo in {"bajo", "verde"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
    elif riesgo in {"sin_datos", "sin datos", "pendiente"}:
        cell.fill = PatternFill("solid", fgColor=COLOR_GRAY)


def agregar_filtros_y_congelar(ws, min_row: int = 1) -> None:
    ws.freeze_panes = f"A{min_row + 1}"
    ws.auto_filter.ref = ws.dimensions


def crear_hoja_sin_datos(ws, mensaje: str) -> None:
    ws.append([mensaje])
    ws["A1"].font = Font(italic=True, color="64748B")


def agregar_tabla(ws, encabezados: list[str], filas: list[list[Any]], sin_datos: str = "Sin datos disponibles") -> None:
    ws.append(encabezados)
    if filas:
        for fila in filas:
            ws.append([valor_excel(valor) for valor in fila])
    else:
        ws.append([sin_datos] + [""] * (len(encabezados) - 1))

    aplicar_estilo_tabla(ws, 1, ws.max_row, 1, len(encabezados))
    agregar_filtros_y_congelar(ws)
    ajustar_ancho_columnas(ws)
    ws.oddFooter.center.text = FOOTER_TEXT


def agregar_kpi(ws, fila: int, columna: int, titulo: str, valor: Any, color: str) -> None:
    ws.cell(fila, columna, titulo)
    ws.cell(fila + 1, columna, valor_excel(valor))
    for cell in (ws.cell(fila, columna), ws.cell(fila + 1, columna)):
        cell.fill = PatternFill("solid", fgColor=color)
        cell.border = Border(
            left=Side(style="thin", color=COLOR_BORDER),
            right=Side(style="thin", color=COLOR_BORDER),
            top=Side(style="thin", color=COLOR_BORDER),
            bottom=Side(style="thin", color=COLOR_BORDER),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(fila, columna).font = Font(bold=True, color="0F172A")
    ws.cell(fila + 1, columna).font = Font(bold=True, color="0F172A", size=14)


def crear_kpi_grid(ws, kpis: list[tuple[str, Any, str]], start_row: int = 7, columns: int = 5) -> int:
    ws.merge_cells(start_row=start_row - 1, start_column=1, end_row=start_row - 1, end_column=10)
    ws.cell(start_row - 1, 1, "Resumen general del sistema")
    ws.cell(start_row - 1, 1).font = Font(bold=True, color="1E3A8A", size=12)
    ws.cell(start_row - 1, 1).fill = PatternFill("solid", fgColor=COLOR_BLUE)
    row = start_row
    col = 1
    for titulo, valor, color in kpis:
        agregar_kpi(ws, row, col, titulo, valor, color)
        col += 2
        if col > columns * 2:
            col = 1
            row += 4
    return row + 3


def agregar_grafico_barras(ws, datos: list[tuple[str, int]], titulo: str, posicion: str) -> None:
    if not datos:
        return
    start_row = ws.max_row + 2
    ws.cell(start_row, 1, "Indicador")
    ws.cell(start_row, 2, "Valor")
    for idx, (label, value) in enumerate(datos, start=start_row + 1):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, int(value or 0))

    chart = BarChart()
    chart.title = titulo
    chart.y_axis.title = "Valor"
    chart.x_axis.title = "Indicador"
    data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(datos))
    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(datos))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, posicion)


def agregar_grafico_pastel(ws, datos: list[tuple[str, int]], titulo: str, posicion: str) -> None:
    if not datos:
        return
    start_row = ws.max_row + 2
    ws.cell(start_row, 1, "Categoría")
    ws.cell(start_row, 2, "Valor")
    for idx, (label, value) in enumerate(datos, start=start_row + 1):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, int(value or 0))

    chart = PieChart()
    chart.title = titulo
    data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(datos))
    labels_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(datos))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.height = 7
    chart.width = 10
    ws.add_chart(chart, posicion)


def aplicar_colores_por_encabezado(ws) -> None:
    encabezados = {cell.value: cell.column for cell in ws[1]}
    for nombre in ("Estado", "Color"):
        col = encabezados.get(nombre)
        if col:
            for column in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
                for cell in column:
                    colorear_estado(cell)
    for nombre in ("Prioridad", "Nivel"):
        col = encabezados.get(nombre)
        if col:
            for column in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
                for cell in column:
                    colorear_prioridad(cell)
    for nombre in ("Nivel de riesgo", "Riesgo IA", "Nivel de validez"):
        col = encabezados.get(nombre)
        if col:
            for column in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
                for cell in column:
                    colorear_riesgo(cell)


def interpretacion_integral(resumen: dict, metricas_data: dict) -> list[str]:
    resumen_metricas = metricas_data.get("resumen_general", {})
    textos = []
    if int(resumen.get("total_alertas_criticas") or resumen_metricas.get("alertas_criticas") or 0) > 0:
        textos.append("Existen alertas críticas que requieren atención prioritaria.")
    if int(resumen_metricas.get("evidencias_sin_archivo") or 0) > 0:
        textos.append("Debe completarse la carga documental de evidencias sin sustento.")
    if int(resumen.get("acciones_pendientes") or resumen_metricas.get("acciones_pendientes") or 0) > 0:
        textos.append("Se recomienda priorizar el seguimiento de acciones de mejora pendientes.")
    if int(resumen.get("total_analisis_ia") or resumen_metricas.get("total_analisis_ia") or 0) > 0:
        textos.append("El sistema cuenta con apoyo de agentes IA para la toma de decisiones.")
    return textos


def interpretacion_silabos(resumen: dict) -> list[str]:
    textos = []
    if int(resumen.get("silabos_pendientes") or 0) > 0:
        textos.append("Existen sílabos pendientes que requieren revisión y actualización.")
    if int(resumen.get("silabos_observados") or 0) > 0:
        textos.append("Los sílabos observados deben atenderse antes del cierre del periodo académico.")
    if int(resumen.get("total_brechas_curriculares") or 0) > 0:
        textos.append("Las brechas curriculares identificadas requieren acciones de mejora asociadas.")
    if int(resumen.get("total_analisis_ia") or 0) > 0:
        textos.append("El módulo cuenta con análisis IA para fortalecer la gestión curricular.")
    return textos


def crear_reporte_silabos_excel(
    resumen: dict,
    silabos: list[dict],
    analisis: list[dict],
    trazabilidad: list[dict],
    brechas: list[dict],
    acciones: list[dict],
) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Resumen Ejecutivo")
    crear_titulo_reporte(ws, "REPORTE DE GESTIÓN DE SÍLABOS", resumen.get("fecha_generacion", ""))
    kpis = [
        ("Total de sílabos", resumen.get("total_silabos"), COLOR_BLUE),
        ("Completos", resumen.get("silabos_completos"), COLOR_GREEN),
        ("Incompletos", resumen.get("silabos_incompletos"), COLOR_RED),
        ("Observados", resumen.get("silabos_observados"), COLOR_YELLOW),
        ("Pendientes", resumen.get("silabos_pendientes"), COLOR_RED),
        ("Cumplimiento", f"{resumen.get('cumplimiento_promedio', 0)}%", COLOR_GREEN),
        ("Análisis IA", resumen.get("total_analisis_ia"), COLOR_PURPLE),
        ("Brechas", resumen.get("total_brechas_curriculares"), COLOR_YELLOW),
        ("Acciones", resumen.get("total_acciones_mejora"), COLOR_BLUE),
    ]
    siguiente_fila = crear_kpi_grid(ws, kpis, start_row=7, columns=5)
    agregar_interpretacion(ws, siguiente_fila + 1, interpretacion_silabos(resumen))
    ajustar_ancho_columnas(ws)
    ws.oddFooter.center.text = FOOTER_TEXT

    ws_silabos = wb.create_sheet("Sílabos")
    agregar_tabla(
        ws_silabos,
        ["ID", "Semestre académico", "Facultad", "Programa de estudios", "Asignatura", "Código", "Ciclo", "Modalidad", "Créditos", "Horas semestrales", "Horas semanales", "Docente", "Correo", "Estado", "Porcentaje de cumplimiento", "Observación", "Archivo URL", "Fecha de registro", "Fecha de actualización"],
        [[item.get("id"), item.get("semestre_academico"), item.get("facultad"), item.get("programa_estudios"), item.get("asignatura"), item.get("codigo_asignatura"), item.get("ciclo"), item.get("modalidad"), item.get("creditos"), item.get("total_horas_semestrales"), item.get("total_horas_semanales"), item.get("docente_responsable"), item.get("correo_docente"), item.get("estado"), item.get("porcentaje_cumplimiento"), item.get("observacion_general"), item.get("archivo_url"), item.get("created_at"), item.get("updated_at")] for item in silabos],
    )
    aplicar_colores_por_encabezado(ws_silabos)

    ws_analisis = wb.create_sheet("Análisis IA")
    agregar_tabla(
        ws_analisis,
        ["ID", "Sílabo ID", "Asignatura", "Modelo usado", "Nivel de cumplimiento", "Resumen", "Recomendaciones", "Fecha de análisis"],
        [[item.get("id"), item.get("silabo_id"), item.get("asignatura"), item.get("modelo_usado"), item.get("nivel_cumplimiento") or item.get("nivel_riesgo"), item.get("resumen"), item.get("recomendaciones") or item.get("sugerencias"), item.get("created_at")] for item in analisis],
        "Sin análisis IA registrados.",
    )

    ws_traz = wb.create_sheet("Trazabilidad Curricular")
    agregar_tabla(
        ws_traz,
        ["ID", "Curso origen", "Curso destino", "Ciclo origen", "Ciclo destino", "Tipo de relación", "Nivel de coherencia", "Observación", "Recomendación", "Fecha"],
        [[item.get("id"), item.get("asignatura_origen") or item.get("curso_origen"), item.get("asignatura_destino") or item.get("curso_destino"), item.get("ciclo_origen"), item.get("ciclo_destino"), item.get("tipo_relacion"), item.get("nivel_coherencia"), item.get("observacion"), item.get("sugerencia") or item.get("recomendacion"), item.get("created_at")] for item in trazabilidad],
        "Sin trazabilidad curricular registrada.",
    )

    ws_brechas = wb.create_sheet("Brechas Curriculares")
    agregar_tabla(
        ws_brechas,
        ["ID", "Tipo de brecha", "Descripción", "Prioridad", "Curso relacionado", "Ciclo", "Recomendación", "Estado", "Fecha"],
        [[item.get("id"), item.get("tipo_brecha"), item.get("descripcion"), item.get("prioridad"), item.get("asignatura") or item.get("curso_relacionado"), item.get("ciclo"), item.get("recomendacion"), item.get("estado"), item.get("created_at")] for item in brechas],
    )
    aplicar_colores_por_encabezado(ws_brechas)

    ws_acciones = wb.create_sheet("Acciones de Mejora")
    agregar_tabla(
        ws_acciones,
        ["ID", "Título", "Descripción", "Prioridad", "Estado", "Responsable", "Origen", "Fecha programada", "Fecha de creación"],
        [[item.get("id"), item.get("titulo"), item.get("descripcion"), item.get("prioridad"), item.get("estado"), item.get("responsable"), item.get("origen_tipo"), item.get("fecha_limite") or item.get("fecha_programada"), item.get("created_at")] for item in acciones],
    )
    aplicar_colores_por_encabezado(ws_acciones)

    ws_graficos = wb.create_sheet("Gráficos")
    aplicar_estilo_titulo(ws_graficos, "A1:H1", "GRÁFICOS DEL REPORTE DE SÍLABOS")
    agregar_grafico_pastel(ws_graficos, [("Completos", resumen.get("silabos_completos", 0)), ("Pendientes", resumen.get("silabos_pendientes", 0)), ("Observados", resumen.get("silabos_observados", 0)), ("Incompletos", resumen.get("silabos_incompletos", 0))], "Distribución de sílabos por estado", "A4")
    agregar_grafico_barras(ws_graficos, [("Total de sílabos", resumen.get("total_silabos", 0)), ("Análisis IA", resumen.get("total_analisis_ia", 0)), ("Brechas curriculares", resumen.get("total_brechas_curriculares", 0)), ("Acciones de mejora", resumen.get("total_acciones_mejora", 0))], "Indicadores principales", "L4")
    acciones_estado = {
        "Pendientes": sum(1 for item in acciones if item.get("estado") == "pendiente"),
        "En proceso": sum(1 for item in acciones if item.get("estado") == "en_proceso"),
        "Completadas": sum(1 for item in acciones if item.get("estado") in {"atendida", "completada", "completado"}),
    }
    agregar_grafico_barras(ws_graficos, list(acciones_estado.items()), "Acciones de mejora por estado", "A22")

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return archivo


def crear_reporte_integral_excel(reporte: dict, metricas: dict) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    resumen = reporte.get("resumen", {})
    metricas_data = metricas.get("data", metricas)

    ws = wb.create_sheet("Resumen Ejecutivo")
    crear_titulo_reporte(ws, "REPORTE INTEGRAL DE MEJORA CONTINUA", reporte.get("fecha_generacion", ""))
    resumen_metricas = metricas_data.get("resumen_general", {})
    kpis = [
        ("Macroprocesos", resumen.get("total_macroprocesos"), COLOR_BLUE),
        ("Evidencias", resumen.get("total_evidencias"), COLOR_BLUE),
        ("Con sustento", resumen_metricas.get("evidencias_con_archivo", 0), COLOR_GREEN),
        ("Sin sustento", resumen_metricas.get("evidencias_sin_archivo", 0), COLOR_RED),
        ("Alertas activas", resumen.get("total_alertas_activas"), COLOR_YELLOW),
        ("Alertas críticas", resumen.get("total_alertas_criticas"), COLOR_RED),
        ("Acciones", resumen.get("total_acciones_mejora"), COLOR_BLUE),
        ("Pendientes", resumen.get("acciones_pendientes"), COLOR_YELLOW),
        ("Análisis IA", resumen.get("total_analisis_ia"), COLOR_PURPLE),
        ("Validaciones IA", resumen.get("total_validaciones_ia"), COLOR_PURPLE),
    ]
    siguiente_fila = crear_kpi_grid(ws, kpis, start_row=7, columns=5)
    agregar_interpretacion(ws, siguiente_fila + 1, interpretacion_integral(resumen, metricas_data))
    ajustar_ancho_columnas(ws)
    ws.oddFooter.center.text = FOOTER_TEXT

    hojas = [
        ("Semáforo y Avance", ["Macroproceso", "Avance promedio", "Color del semáforo", "Alertas críticas", "Alertas altas", "Riesgo IA", "Mensaje de interpretación"], [[item.get("nombre") or item.get("macroproceso"), item.get("avance_promedio"), item.get("color"), item.get("alertas_criticas"), item.get("alertas_altas"), item.get("riesgo_ia"), item.get("mensaje")] for item in reporte.get("semaforo", [])]),
        ("Evidencias", ["Código", "Macroproceso", "Título", "Tipo de evidencia", "Responsable", "Mes programado", "Estado", "Prioridad", "Avance", "Archivo de sustento", "Observación", "Origen documento", "Fecha programada", "Fecha cumplimiento"], [[item.get("codigo"), item.get("macroproceso"), item.get("titulo"), item.get("tipo_evidencia"), item.get("responsable"), item.get("mes_programado"), item.get("estado"), item.get("prioridad"), item.get("avance"), item.get("archivo_url"), item.get("observacion"), item.get("origen_documento"), item.get("fecha_programada"), item.get("fecha_cumplimiento")] for item in reporte.get("evidencias", reporte.get("evidencias_criticas", []))]),
        ("Seguimiento semanal", ["Macroproceso", "Código evidencia", "Título evidencia", "Semana inicio", "Semana fin", "Responsable", "Nivel avance", "Porcentaje avance", "Acción realizada", "Resultado observado", "Dificultad encontrada", "Compromiso siguiente", "Requiere apoyo", "Tipo apoyo", "Archivo sustento URL", "Fecha de registro"], [[item.get("macroproceso"), item.get("codigo_evidencia"), item.get("titulo_evidencia"), item.get("semana_inicio"), item.get("semana_fin"), item.get("responsable"), item.get("nivel_avance"), item.get("porcentaje_avance"), item.get("descripcion_accion"), item.get("resultado_observado"), item.get("dificultad_encontrada"), item.get("compromiso_siguiente_semana"), item.get("requiere_apoyo"), item.get("tipo_apoyo_requerido"), item.get("archivo_sustento_url"), item.get("created_at")] for item in reporte.get("seguimientos_semanales", [])]),
        ("Alertas", ["Macroproceso", "Código", "Título", "Descripción", "Nivel", "Estado", "Recomendación", "Fecha creación"], [[item.get("macroproceso"), item.get("codigo"), item.get("titulo"), item.get("descripcion"), item.get("nivel_alerta"), item.get("estado"), item.get("recomendacion"), item.get("created_at")] for item in reporte.get("alertas_activas", [])]),
        ("Acciones de Mejora", ["Título", "Descripción", "Macroproceso", "Responsable", "Prioridad", "Estado", "Origen", "Fecha programada", "Fecha creación"], [[item.get("titulo"), item.get("descripcion"), item.get("macroproceso"), item.get("responsable"), item.get("prioridad"), item.get("estado"), item.get("origen_tipo"), item.get("fecha_limite") or item.get("fecha_programada"), item.get("created_at")] for item in reporte.get("acciones_mejora", [])]),
        ("Análisis IA", ["Macroproceso", "Tipo de análisis", "Nivel de riesgo", "Resumen", "Modelo usado", "Fecha"], [[item.get("macroproceso"), item.get("tipo_analisis"), item.get("nivel_riesgo"), item.get("resumen"), item.get("modelo_usado"), item.get("created_at")] for item in reporte.get("ultimos_analisis_ia", [])]),
        ("Validaciones Documentales", ["Código evidencia", "Macroproceso", "Nivel de validez", "Pertinencia", "Resumen", "Elementos detectados", "Elementos faltantes", "Recomendaciones", "Modelo usado", "Fecha"], [[item.get("evidencia_id"), item.get("macroproceso"), item.get("nivel_validez"), item.get("pertinencia"), item.get("resumen"), item.get("elementos_detectados"), item.get("elementos_faltantes"), item.get("recomendaciones"), item.get("modelo_usado"), item.get("created_at")] for item in reporte.get("validaciones_documentales", [])]),
        ("Métricas", ["Indicador", "Valor", "Interpretación"], [[item.get("indicador"), item.get("valor"), item.get("interpretacion")] for item in metricas_data.get("indicadores_clave", [])]),
    ]
    for titulo, encabezados, filas in hojas:
        ws_tabla = wb.create_sheet(titulo)
        agregar_tabla(ws_tabla, encabezados, filas)
        aplicar_colores_por_encabezado(ws_tabla)

    if "Semáforo y Avance" in wb.sheetnames:
        ws_semaforo = wb["Semáforo y Avance"]
        agregar_grafico_barras(
            ws_semaforo,
            [(item.get("nombre") or item.get("macroproceso") or "-", item.get("avance_promedio") or 0) for item in reporte.get("semaforo", [])],
            "Avance promedio por macroproceso",
            "I3",
        )

    if "Análisis IA" in wb.sheetnames:
        ws_ia = wb["Análisis IA"]
        metricas_ia = metricas_data.get("metricas_ia", {})
        ws_ia["H1"] = "Uso de agentes IA"
        ws_ia["H1"].font = Font(bold=True, color="FFFFFF")
        ws_ia["H1"].fill = PatternFill("solid", fgColor=COLOR_DARK)
        ia_kpis = [
            ("Total análisis IA", metricas_ia.get("analisis_ia_ejecutados", resumen.get("total_analisis_ia", 0))),
            ("Análisis planificación", metricas_ia.get("analisis_planificacion", 0)),
            ("Análisis gestión académica", metricas_ia.get("analisis_gestion_academica", 0)),
            ("Análisis coordinador", metricas_ia.get("analisis_integral_mejora_continua", 0)),
            ("Análisis sílabos", metricas_ia.get("analisis_silabos", 0)),
            ("Validaciones IA", metricas_ia.get("validaciones_documentales_ia", resumen.get("total_validaciones_ia", 0))),
        ]
        for index, (label, value) in enumerate(ia_kpis, start=2):
            ws_ia.cell(index, 8, label)
            ws_ia.cell(index, 9, value)
            ws_ia.cell(index, 8).font = Font(bold=True, color="334155")
            ws_ia.cell(index, 9).fill = PatternFill("solid", fgColor=COLOR_PURPLE)
            ws_ia.cell(index, 9).alignment = Alignment(horizontal="center")
        ajustar_ancho_columnas(ws_ia)

    ws_graficos = wb.create_sheet("Gráficos")
    aplicar_estilo_titulo(ws_graficos, "A1:H1", "GRÁFICOS DEL REPORTE INTEGRAL")
    semaforo = reporte.get("semaforo", [])
    agregar_grafico_barras(ws_graficos, [(item.get("nombre") or item.get("macroproceso") or "-", item.get("avance_promedio") or 0) for item in semaforo], "Avance por macroproceso", "A4")
    evidencias = reporte.get("evidencias", reporte.get("evidencias_criticas", []))
    resumen_metricas = metricas_data.get("resumen_general", {})
    sustento = [
        ("Con sustento", resumen_metricas.get("evidencias_con_archivo", sum(1 for item in evidencias if item.get("archivo_url")))),
        ("Sin sustento", resumen_metricas.get("evidencias_sin_archivo", sum(1 for item in evidencias if not item.get("archivo_url")))),
    ]
    agregar_grafico_pastel(ws_graficos, sustento, "Evidencias con sustento vs sin sustento", "A18")
    alertas = reporte.get("alertas_activas", [])
    alertas_nivel = {nivel: sum(1 for item in alertas if item.get("nivel_alerta") == nivel) for nivel in ["critica", "alta", "media", "baja"]}
    agregar_grafico_barras(ws_graficos, list(alertas_nivel.items()), "Alertas por nivel", "J4")
    acciones = reporte.get("acciones_mejora", [])
    acciones_estado = {estado: sum(1 for item in acciones if item.get("estado") == estado) for estado in ["pendiente", "en_proceso", "atendida", "descartada"]}
    agregar_grafico_barras(ws_graficos, list(acciones_estado.items()), "Acciones por estado", "J18")
    metricas_ia = metricas_data.get("metricas_ia", {})
    agregar_grafico_barras(
        ws_graficos,
        [
            ("Planificación", metricas_ia.get("analisis_planificacion", 0)),
            ("Gestión académica", metricas_ia.get("analisis_gestion_academica", 0)),
            ("Coordinador", metricas_ia.get("analisis_integral_mejora_continua", 0)),
            ("Sílabos", metricas_ia.get("analisis_silabos", 0)),
            ("Validaciones IA", metricas_ia.get("validaciones_documentales_ia", resumen.get("total_validaciones_ia", 0))),
        ],
        "Uso de IA por tipo de agente",
        "A35",
    )

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return archivo
